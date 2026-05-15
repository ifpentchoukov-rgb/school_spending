"""New York extractor — NYSED ST-3 Annual Financial Report bulk XLSX.

Source: https://stateaid.nysed.gov/st3/st3data.htm
File pattern: https://stateaid.nysed.gov/st3/st3data/
              {fy-1}-{fy}_School_Year_{fy}-{fy+1}_SAMS%20ST-3.xlsx
Example:      2023-2024_School_Year_2024-2025_SAMS%20ST-3.xlsx
              covers FY24 (school year ending June 30, 2024).

NYSED publishes the ST-3 about 12 months after FY end (FY24 data was
frozen 2025-06-11 per the file's metadata). Posted in "SAMS"
(State Aid Management System) cycle for the following school year.

What this gives us:
  - Per-LEA General Fund operating expenditures for ~698 NY school
    districts. Aligned with the F-33 'current expenditures' frame and
    the actuals topline used for TX/CA/FL/IL/GA/OH/MI/etc.
  - The same XLSX has every other ST-3 line item (Special Aid Fund,
    Capital Fund, Debt Service, Trust & Agency, etc.) — see SFList
    sheet for the full ~4,200-column legend.

Topline definition:
  ST-3 Data sheet, column where row 1 (Legacy) == '49:459' — Account
  Code 'AT9999.0', Schedule A4c Line 463, "TOTAL GENERAL FUND
  EXPENDITURES AND INTERFUND TRANSFERS — Actual Column". Includes
  instruction + admin + transportation + plant O&M + employee
  benefits + debt service + interfund transfers within the General
  Fund. Excludes Special Aid Fund (federal), Capital Fund, Debt
  Service Fund, etc. — those are reported under separate fund
  schedules. Most NY districts run nearly all operating spend through
  the General Fund, so this is a reasonable F-33-aligned topline.

Status: `actual` — post-AFR audited.

NY school FY = July 1 – June 30 (state default). FY24 = SY 2023-24.
Latest publication is FY24; FY25 expected mid-2026.

Crosswalk:
  Master state_leaid format: 'NY-{12-digit BEDS code}'
                              (e.g. 'NY-010100010000' Albany City SD)
  ST-3 BEDS code:            6-digit string (e.g. '010100')
  → strip 'NY-' prefix; first 6 chars of suffix == ST-3 BEDS.

Known gaps:
  - Charter schools file ST-3D (a separate form); they are NOT in this
    file. NY has ~360 charter LEAs in the master that this extractor
    will skip — they need a sibling extractor against the ST-3D
    publication when discovered.
  - BOCES are not in the master operating-LEA universe, so they're
    correctly skipped here too.
"""

from __future__ import annotations

import argparse
import sys
import urllib.error
import urllib.request
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)
from extractors._exceptions import SourceNotYetPublished

EXTRACTOR_NAME = "ny"
STATE = "NY"
BUCKET = "ny"
SOURCE_PORTAL_URL = "https://stateaid.nysed.gov/st3/st3data.htm"
PUBLISHER = "New York State Education Department (Office of State Aid)"
DOCUMENT_TYPE = "nysed_st3_xlsx"
TOPLINE_DEFINITION = (
    "NYSED ST-3 'ST-3 Data' sheet, AT9999.0 (Schedule A4c Line 463) Actual "
    "Column = TOTAL GENERAL FUND EXPENDITURES AND INTERFUND TRANSFERS. "
    "General-Fund-only; aligned with F-33 'current expenditures' frame."
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

# Identifies the AT9999.0 (Total GF Expenditures, Actual) column in the
# ST-3 Data sheet. The header at row 1 contains the Legacy code; we
# scan for this exact string. NYSED's account-code → legacy-code map
# is in SFList; this entry is row 461 there with Account Code AT9999.0.
TARGET_LEGACY_CODE = "49:459"

# Phase 7.4 — canonical-category mapping for NY ST-3.
# Each entry: canonical_category -> (legacy_code, definition_fragment).
# Code structure:
#   AT*.0 within the GF (49:XXX rows) are GF function subtotals.
#   HT9999, FT9999, CT9999, VT9999 are cross-fund totals (Capital,
#   Special Aid, School Lunch, Debt Service Funds).
# 8 categories emitted; gaps:
#   support_services_instruction — not a clean NY bucket (Schedule A4b
#     "Administration and Improvement" mixes admin + curriculum).
#   operations_maintenance — folded into NY Schedule A4a General Support
#     (A1620-A1639); to break out we'd need to sum individual A16XX
#     codes. Deferred; for now O&M is part of `administration`.
#   revenue_federal/state/local — Schedule A3 (GF Revenues) codes
#     not yet mapped. Deferred.
NY_COMPONENT_MAPPING: dict[str, tuple[str, str]] = {
    "instruction": (
        "49:301",
        "ST-3 Schedule A4b Line 326 AT2999.0 TOTAL INSTRUCTION — Actual Column",
    ),
    "support_services_student": (
        "49:300",
        "ST-3 Schedule A4b Line 325 AT2899.0 TOTAL PUPIL SERVICES — Actual Column "
        "(includes guidance, health, psychology, co-curricular)",
    ),
    "administration": (
        "49:124",
        "ST-3 Schedule A4a Line 135 AT1999.0 TOTAL GENERAL SUPPORT — Actual Column "
        "(includes Board of Education, Central Admin, Finance, Staff, Central "
        "Services — and operations/maintenance, which NY does not break out "
        "separately in this aggregate)",
    ),
    "transportation": (
        "49:330",
        "ST-3 Schedule A4c Line 345 AT5599.0 TOTAL PUPIL TRANSPORTATION — Actual Column",
    ),
    "employee_benefits": (
        "49:380",
        "ST-3 Schedule A4c Line 384 AT9098.0 TOTAL EMPLOYEE BENEFITS — Actual Column "
        "(General Fund only — health/dental/retirement/FICA across all "
        "function categories)",
    ),
    "debt_service": (
        "49:451",
        "ST-3 Schedule A4c Line 456 AT9898.0 TOTAL DEBT SERVICE — Actual Column "
        "(General Fund only; Debt Service Fund VT9999 is separate)",
    ),
    "food_service": (
        "59:019",
        "ST-3 Schedule C3 Line 19 CT9999.0 TOTAL SCHOOL FOOD SERVICE PROGRAMS "
        "EXPENDITURES AND INTERFUND TRANSFERS — Actual Column "
        "(School Lunch Fund — separate from General Fund)",
    ),
    "capital_outlay": (
        "67:171",
        "ST-3 Schedule G3 Line 14 HT9999.0 TOTAL CAPITAL FUND EXPENDITURES AND "
        "INTERFUND TRANSFERS — Actual Column (Capital Fund — separate from "
        "General Fund; not double-counted in topline)",
    ),
}

# Layout of ST-3 Data (1-indexed by openpyxl, 0-indexed in iter_rows):
#   row 1 (idx 0): column index numbers (1, 2, 3, ...)
#   row 2 (idx 1): Legacy codes per column ('45:001', '45:002', ...)
#   row 3 (idx 2): RefKey codes per column
#   row 4 (idx 3): Line numbers
#   rows 5-13: misc metadata
#   row 14 (idx 13): 'Beds Code' / 'District Name' header for cols A-B
#   rows 15+ (idx 14+): one entity per row — col A = BEDS, col B = name,
#     cols E+ = financial values per ST-3 line item
LEGACY_HEADER_ROW_IDX = 1
ENTITIES_START_ROW_IDX = 14
BEDS_COL_IDX = 0
NAME_COL_IDX = 1


def _fy_label(fy: int) -> str:
    """fy=2024 → '2023-2024'."""
    return f"{fy - 1:04d}-{fy:04d}"


def file_url(fiscal_year: int) -> str:
    sams_cycle = _fy_label(fiscal_year + 1)
    return (
        "https://stateaid.nysed.gov/st3/st3data/"
        f"{_fy_label(fiscal_year)}_School_Year_"
        f"{sams_cycle}_SAMS%20ST-3.xlsx"
    )


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        if e.code == 404:
            raise SourceNotYetPublished(
                f"NYSED ST-3 404 at {url} — FY data publishes ~12 months "
                "after FY end; not yet posted."
            ) from e
        raise


def parse_st3(xlsx_bytes: bytes) -> list[dict]:
    """Read the ST-3 Data sheet and pull AT9999.0 (Total GF Exp Actual)
    plus the Phase 7.4 canonical category columns per BEDS code.
    Returns [{beds, name, total_gf_exp, components: {category: amount}}, ...]."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "ST-3 Data" not in wb.sheetnames:
        raise RuntimeError(
            f"'ST-3 Data' sheet not found; sheets={wb.sheetnames}"
        )
    ws = wb["ST-3 Data"]

    target_col_idx: int | None = None
    # category -> column index for that category's Legacy code
    component_col_idx: dict[str, int] = {}
    out: list[dict] = []
    needed_codes: set[str] = {TARGET_LEGACY_CODE} | {
        code for code, _def in NY_COMPONENT_MAPPING.values()
    }
    code_to_category: dict[str, str] = {
        code: cat for cat, (code, _) in NY_COMPONENT_MAPPING.items()
    }

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == LEGACY_HEADER_ROW_IDX:
            for j, v in enumerate(row):
                if v not in needed_codes:
                    continue
                if v == TARGET_LEGACY_CODE:
                    target_col_idx = j
                else:
                    component_col_idx[code_to_category[v]] = j
            if target_col_idx is None:
                raise RuntimeError(
                    f"Could not find target Legacy code '{TARGET_LEGACY_CODE}' "
                    f"(AT9999.0 Total GF Exp Actual) in row {LEGACY_HEADER_ROW_IDX}"
                )
        elif i >= ENTITIES_START_ROW_IDX and target_col_idx is not None:
            beds_raw = row[BEDS_COL_IDX]
            name_raw = row[NAME_COL_IDX]
            value_raw = row[target_col_idx]
            if not beds_raw or value_raw is None:
                continue
            try:
                value = float(value_raw)
            except (TypeError, ValueError):
                continue
            if value <= 0:
                continue  # don't emit zero/negative toplines

            # Pull canonical-category amounts for this BEDS row.
            components: dict[str, float] = {}
            for category, col_idx in component_col_idx.items():
                cv = row[col_idx]
                if cv is None:
                    continue
                try:
                    components[category] = float(cv)
                except (TypeError, ValueError):
                    continue

            out.append({
                "beds": str(beds_raw).strip().zfill(6),
                "name": str(name_raw or "").strip(),
                "total_gf_exp": value,
                "components": components,
            })
    return out


def build_ny_crosswalk(client: Client) -> dict[str, dict]:
    """Build {6-digit-BEDS: master district row}. Master state_leaid
    is 'NY-{6-digit BEDS}{2-digit type}{4-digit code}'; first 6 digits ==
    BEDS. Type code 86 = charter school. Charters file ST-3D (separate
    form) and aren't in the ST-3 file we're parsing — but their
    state_leaid suffix often shares the first 6 chars with a non-charter
    district (e.g. ROCHESTER ACADEMY OF SCIENCE CHARTER 'NY-261600861193'
    shares its BEDS6 with ROCHESTER CITY SD 'NY-261600010000'). We skip
    charters here; the BEDS6 → district map is then 1:1 with the
    non-charter universe ST-3 covers."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if not sl.startswith("NY-") or len(sl) < 11:
            continue
        suffix = sl[3:]                  # 'XXXXXXTTYYYY'
        beds6 = suffix[:6]               # 'XXXXXX'
        type_code = suffix[6:8]          # 'TT'
        if type_code == "86":
            continue                     # skip charters; they file ST-3D
        out[beds6] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"NY extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        print(f"  downloading {url.rsplit('/', 1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.1f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/sams_st3.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'ST-3 Data'; column where row 2 (1-indexed) == "
                f"'{TARGET_LEGACY_CODE}' (AT9999.0 Total GF Exp Actual); "
                "rows 15+ are entities with col A=BEDS, col B=name; "
                "match BEDS == first 6 digits of state_leaid suffix"
            ),
            notes=f"FY{fiscal_year} NYSED ST-3 — General Fund total expenditures",
        )

        crosswalk = build_ny_crosswalk(client)
        print(f"  NY crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_totals = parse_st3(xlsx_bytes)
        print(f"  ST-3 entities with positive GF Exp Actual: {len(district_totals):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for d in district_totals:
            district = crosswalk.get(d["beds"])
            if district is None:
                no_match.append(d["beds"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_gf_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            # Phase 7.4 — emit canonical category components.
            components: list[ComponentInput] = []
            for category, amount in d.get("components", {}).items():
                if amount is None:
                    continue
                _legacy, definition = NY_COMPONENT_MAPPING[category]
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=definition,
                        line_or_cell_reference=(
                            f"Sheet 'ST-3 Data'; column where row 2 == "
                            f"'{_legacy}'; row for BEDS=={d['beds']}"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched BEDS: {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024,
                   help="ST-3 FY (latest as of 2026-05-14: 2024)")
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
