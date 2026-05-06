"""Idaho extractor — ISDE 20-Year Revenues & Expenditures workbook.

Source: https://www.sde.idaho.gov/finance-transparency/public-school-finance/
File: 2004-2024-Revenues-and-Expenditures.xlsx — ISDE Public School
Finance Division publishes a single multi-FY workbook with one sheet
per FY × {M&O Fund, All Funds} × {Revenues, Expenditures}.

What this gives us:
  - Per-district expenditure breakdown across All Funds for ~230
    Idaho LEAs (137 operating + charter LEAs). Each FY has an
    "All Funds Expd & by ADA" sheet with columns:
    Instruction, Support Services, Non-Instructional, Capital Assets,
    Debt Services, Total.

Topline definition:
  Sheet 'FY{YYYY} All Funds Expd & by ADA' — sum of cols
  2 (Instruction) + 3 (Support Services) + 4 (Non-Instructional) per
  district. Excludes Capital Assets (col 5) and Debt Services (col 6).
  Aligned with F-33 'current expenditures' frame.

Status: `actual` — post-AFR audited.

Crosswalk:
  Master state_leaid format: 'ID-{3-digit}' (e.g. 'ID-001' Boise)
  XLSX district number:      integer (e.g. 1)
  → state_leaid suffix == zfill(district_number, 3).
"""

from __future__ import annotations

import argparse
import io
import sys
import urllib.error
import urllib.request

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "id"
STATE = "ID"
BUCKET = "id"
SOURCE_PORTAL_URL = "https://www.sde.idaho.gov/finance-transparency/public-school-finance/"
PUBLISHER = "Idaho State Department of Education (Public School Finance Division)"
DOCUMENT_TYPE = "isde_revenues_expenditures_xlsx"
TOPLINE_DEFINITION = (
    "ISDE 20-Year Revenues & Expenditures workbook, sheet 'FY{YYYY} "
    "All Funds Expd & by ADA' — sum of Instruction + Support Services "
    "+ Non-Instructional per district. Excludes Capital Assets and "
    "Debt Services. Aligned with F-33 'current expenditures' frame."
)
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

KNOWN_FILE_URLS: dict[int, str] = {
    # 2004-2024 multi-FY file; FY24 is the latest sheet.
    2024: "https://www.sde.idaho.gov/wp-content/uploads/2025/10/2004-2024-Revenues-and-Expenditures.xlsx",
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def parse_id(xlsx_bytes: bytes, fiscal_year: int) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    # Sheet name pattern varies slightly: "FY{YYYY} All Funds Expd & by ADA"
    # for 2024, but "FY {YYYY}" (with space) for 2014-2019 — start broad.
    target = None
    for name in wb.sheetnames:
        if (str(fiscal_year) in name and "All Funds Expd" in name):
            target = name
            break
    if target is None:
        raise RuntimeError(
            f"No 'All Funds Expd' sheet for FY{fiscal_year}; sheets={wb.sheetnames[:6]}"
        )
    ws = wb[target]
    out: list[dict] = []
    rows = ws.iter_rows(values_only=True)
    # Skip 3 title rows; row index 2 has 'School District / Charter School' header
    for r in rows:
        if not r:
            continue
        # Data rows start when col 0 is an int (district number)
        if not isinstance(r[0], (int, float)):
            continue
        try:
            d = int(r[0])
        except (TypeError, ValueError):
            continue
        # Sum cols 2 (Instruction), 3 (Support Services), 4 (Non-Instructional)
        total = 0.0
        for c in (2, 3, 4):
            v = r[c] if c < len(r) else None
            if v is None:
                continue
            try:
                total += float(v)
            except (TypeError, ValueError):
                continue
        if total <= 0:
            continue
        out.append({"code": f"{d:03d}", "total_op_exp": total})
    return out


def build_id_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("ID-"):
            out[sl.removeprefix("ID-")] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"ID extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No ISDE R&E URL for fiscal_year={fiscal_year}; add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/', 1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.2f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/revenues_expenditures.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'FY{YYYY} All Funds Expd & by ADA'; sum cols 2 "
                "(Instruction) + 3 (Support Services) + 4 (Non-Instructional) "
                "per row; match zfill(col0, 3) == state_leaid suffix"
            ),
            notes=f"FY{fiscal_year} ISDE 20-Year R&E workbook",
        )

        crosswalk = build_id_crosswalk(client)
        print(f"  ID crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_data = parse_id(xlsx_bytes, fiscal_year=fiscal_year)
        print(f"  ISDE districts with FY{fiscal_year} data: {len(district_data):,}")

        no_match: list[str] = []
        for d in district_data:
            district = crosswalk.get(d["code"])
            if district is None:
                no_match.append(d["code"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            _, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched ID codes (charters/specialty): {len(no_match)}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
