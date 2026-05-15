"""Pennsylvania extractor — PDE General Fund Budget (GFB) bulk Excel.

Source: https://www.pa.gov/agencies/education/programs-and-services/schools/
        grants-and-funding/school-finances/financial-data/general-fund-budget-gfb-data
File pattern: https://www.pa.gov/content/dam/copapwp-pagov/en/education/
              documents/schools/grants-and-funding/school-finances/finances/
              gfbdata/{YYYY-YY}gfbdata.xlsx
e.g. 2025-26gfbdata.xlsx covers SY 2025-26 = our fiscal_year=2026.

What this gives us:
  - Adopted General Fund Budget (operating budget) per PA school district per
    fiscal year. PDE publishes one Excel per fiscal year covering all
    ~500 districts plus IUs (Intermediate Units, separate sheets).
  - Available years: 2016-17 through 2025-26 as of 2026-05-05.

Topline definition:
  `FB_Cert` sheet, column `TotalExpAmount` — the certified total
  expenditure budget per district. PDE derives this from the detailed
  expenditure breakdown in the `Exp` sheet (function-object grid). Aligned
  with adopted-budget definition used by FL Summary Budget and CA SACS BS1
  for cross-state comparability.

Status: `adopted` — these are board-adopted budgets filed with PDE per
24 P.S. § 6-687. They're NOT actuals; PA AFR is a separate cycle and
covered by a sibling extractor TBD.

Crosswalk:
  Master state_leaid format: 'PA-{9-digit-AUN}' (e.g. 'PA-101260303')
  GFB AUN column:            9-digit AUN (e.g. 101260303)
  → strip 'PA-'.
"""

from __future__ import annotations

import argparse
import sys
import urllib.request
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "pa"
STATE = "PA"
BUCKET = "pa"
SOURCE_PORTAL_URL = (
    "https://www.pa.gov/agencies/education/programs-and-services/schools/"
    "grants-and-funding/school-finances/financial-data/general-fund-budget-gfb-data"
)
PUBLISHER = "Pennsylvania Department of Education"
DOCUMENT_TYPE = "pde_gfb_xlsx"
TOPLINE_DEFINITION = (
    "PDE General Fund Budget, FB_Cert sheet, TotalExpAmount column "
    "(certified total adopted operating expenditure budget per district)"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"


def _fy_full(fiscal_year: int) -> str:
    """2026 → '2025-26'."""
    return f"{fiscal_year - 1:04d}-{fiscal_year % 100:02d}"


def file_url(fiscal_year: int) -> str:
    return (
        "https://www.pa.gov/content/dam/copapwp-pagov/en/education/documents/"
        "schools/grants-and-funding/school-finances/finances/gfbdata/"
        f"{_fy_full(fiscal_year)}gfbdata.xlsx"
    )


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        if e.code == 404:
            from extractors._exceptions import SourceNotYetPublished
            raise SourceNotYetPublished(
                f"PA GFB 404 at {url} — FY27 GFB certifies ~Sept 2026 per "
                "24 P.S. § 6-687; not yet published."
            ) from e
        raise


def parse_fb_cert(xlsx_bytes: bytes) -> list[dict]:
    """Read FB_Cert sheet, return rows for InstCat='01' (school districts)."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "FB_Cert" not in wb.sheetnames:
        raise RuntimeError(f"FB_Cert sheet not found; sheets={wb.sheetnames}")
    ws = wb["FB_Cert"]
    out: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 6:
            continue
        cat, aun, name, county = row[0], row[1], row[2], row[3]
        if cat != "01" or not aun:
            continue
        try:
            total_exp = float(row[5]) if row[5] is not None else None
        except (TypeError, ValueError):
            total_exp = None
        if total_exp is None or total_exp <= 0:
            continue
        out.append({
            "aun": str(aun),
            "name": name,
            "county": county,
            "total_exp": total_exp,
        })
    return out


# Phase 7.4 — canonical category mapping for PA. The 'Exp' sheet has
# columns named '<function>-<object>' (e.g. '1100-100' = Regular
# Instruction Salaries). PA Chart of Accounts (per 22 Pa. Code §
# 102.32):
#   1100-1499 — Instruction (Regular/Special/Vocational/Other)
#   2100      — Pupil Personnel Services (counsel, social, psych)
#   2200      — Instructional Staff Support (curriculum, library)
#   2300      — Administrative Services (board, exec, fiscal)
#   2400      — Pupil Health
#   2500      — Business Services
#   2600      — Operation & Maintenance of Plant Services
#   2700      — Student Transportation
#   2800/2900 — Central / Other Support Services
#   3100      — Food Services
#   3200      — Student Activities
#   3300      — Community Services
#   4000      — Facilities Acquisition, Construction & Improvement (Capital)
#   5100      — Debt Service
#   5200-5900 — Other Financing Uses
# Objects:
#   100 = Salaries; 200 = Benefits; 300-800 = various purchased/supplies
PA_CATEGORY_FUNCTION_PREFIXES: dict[str, list[str]] = {
    "instruction": ["1100", "1200", "1300", "1400", "1500", "1600", "1700", "1800"],
    "support_services_student": ["2100", "2400"],
    "support_services_instruction": ["2200"],
    "administration": ["2300", "2500", "2800", "2900"],
    "operations_maintenance": ["2600"],
    "transportation": ["2700"],
    "food_service": ["3100"],
    "capital_outlay": ["4000"],
    "debt_service": ["5100"],
}


def parse_exp(xlsx_bytes: bytes) -> dict[str, dict]:
    """Read Exp sheet. Returns {aun: {category: amount, ..., 'employee_benefits': X}}.

    Aggregates the function-object grid into canonical categories.
    Employee benefits = sum of all -200 (Object 200) cells across
    functions, capturing benefits regardless of which function they
    sit in.
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "Exp" not in wb.sheetnames:
        raise RuntimeError(f"Exp sheet not found; sheets={wb.sheetnames}")
    ws = wb["Exp"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = rows[0]
    # Build column index -> (function prefix, object code) mapping
    col_function: dict[int, str] = {}
    col_object: dict[int, str] = {}
    for j, h in enumerate(header[4:], start=4):
        if not h or "-" not in str(h):
            continue
        func, _, obj = str(h).partition("-")
        col_function[j] = func.strip()
        col_object[j] = obj.strip()
    # Reverse-index: function prefix -> categories
    func_to_cats: dict[str, list[str]] = {}
    for cat, prefixes in PA_CATEGORY_FUNCTION_PREFIXES.items():
        for p in prefixes:
            func_to_cats.setdefault(p, []).append(cat)

    out: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        if not row or len(row) < 4:
            continue
        cat, aun, _name, _county = row[0], row[1], row[2], row[3]
        if cat != "01" or not aun:
            continue
        district_totals: dict[str, float] = {c: 0.0 for c in PA_CATEGORY_FUNCTION_PREFIXES}
        district_totals["employee_benefits"] = 0.0
        for j, v in enumerate(row[4:], start=4):
            if v is None:
                continue
            try:
                amt = float(v)
            except (TypeError, ValueError):
                continue
            func = col_function.get(j)
            obj = col_object.get(j)
            if not func:
                continue
            # Function-based aggregation
            for cat_name in func_to_cats.get(func, ()):
                district_totals[cat_name] += amt
            # Object 200 — benefits regardless of function
            if obj == "200":
                district_totals["employee_benefits"] += amt
        # Drop zero categories to keep the upsert payload tight
        out[str(aun)] = {
            cat: amt for cat, amt in district_totals.items() if amt > 0
        }
    return out


def build_pa_crosswalk(client: Client) -> dict[str, dict]:
    """9-digit AUN → district row."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("PA-"):
            out[sl.removeprefix("PA-")] = r
    return out


def extract(*, fiscal_year: int = 2026, triggered_by: str = "manual") -> dict:
    print(f"PA extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        print(f"  downloading {url.rsplit('/',1)[-1]}...")
        try:
            xlsx_bytes = download(url)
        except urllib.error.HTTPError as e:
            print(f"  FAILED: {e}")
            raise
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.1f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/gfb_data.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'FB_Cert'; filter InstCat='01'; "
                "match AUN == state_leaid suffix; topline = TotalExpAmount"
            ),
            notes=(
                f"FY{fiscal_year} adopted GFB; one bulk file covers ~500 PA "
                "school districts. IU (Intermediate Unit) data on separate "
                "sheets; not extracted here."
            ),
        )

        crosswalk = build_pa_crosswalk(client)
        print(f"  PA crosswalk: {len(crosswalk):,} state→NCES mappings")

        gfb_rows = parse_fb_cert(xlsx_bytes)
        print(f"  GFB FB_Cert districts: {len(gfb_rows):,}")

        exp_by_aun = parse_exp(xlsx_bytes)
        print(f"  Exp sheet districts: {len(exp_by_aun):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for row in gfb_rows:
            district = crosswalk.get(row["aun"])
            if district is None:
                no_match.append(row["aun"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="adopted",
                topline_amount=row["total_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            # Phase 7.4 — emit canonical-category components.
            components: list[ComponentInput] = []
            exp_totals = exp_by_aun.get(row["aun"], {})
            for category, amount in exp_totals.items():
                if amount <= 0:
                    continue
                if category == "employee_benefits":
                    definition = (
                        "PA Chart of Accounts: sum of Object 200 (Benefits) "
                        "cells across all functions in Exp sheet for this AUN"
                    )
                else:
                    prefixes = PA_CATEGORY_FUNCTION_PREFIXES[category]
                    definition = (
                        f"PA Chart of Accounts: sum of <fn>-<obj> cells where "
                        f"function prefix in {prefixes} in Exp sheet for this AUN"
                    )
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=definition,
                        line_or_cell_reference=(
                            f"Sheet 'Exp'; aggregate by function prefix; AUN={row['aun']}"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched AUNs: {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )
        if no_match[:5]:
            print(f"  sample unmatched: {no_match[:5]}")

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "components_inserted": n_components_inserted,
        "components_updated": n_components_updated,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2026,
                   help="GFB file FY (latest as of 2026-05-05: 2026 = SY 2025-26)")
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
