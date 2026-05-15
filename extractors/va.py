"""Virginia extractor — APA Comparative Report of Local Government Revenues
and Expenditures, Exhibit C (Education total per locality).

Source: https://www.apa.virginia.gov/local-government/reports?type=comparative-reports
File: APA publishes one Excel per fiscal year (FY ending June 30) covering
all VA cities + counties + towns. The file URL has an opaque GUID that
changes per FY; the catalog page lists the link.

What this gives us:
  - Per-locality (city or county) education expenditures from APA's
    Comparative Report Exhibit C, column "Education (Exhibit C-6)".
    APA aggregates from the locally-filed Comparative Report Transmittal
    each county/city files annually after their CAFR.
  - 38 cities + 95 counties = 133 localities. Towns also reported but
    don't operate schools.

Topline definition:
  Exhibit C, col 22 ('Education (Exhibit C-6)') — total education
  expenditures per locality. This sums Instruction + Administration +
  Pupil Transportation + Operations & Maintenance + other education
  functional categories. Aligned with F-33 'current expenditures' frame.

Status: `actual` — these are post-CAFR audited numbers.

Note: VDOE's Superintendent's Annual Report (the legacy source)
returned 403 from our IP/UA via Akamai. APA's Comparative Report is the
public-API-equivalent for cross-locality school finance comparison.

Crosswalk:
  Master state_leaid format: 'VA-{3-digit-VDOE-code}' (e.g. 'VA-101' Alexandria)
  APA locality column:       Place name (e.g. 'Alexandria' for city,
                             'Fairfax' for county)
  → name match against master lea_name UPPER, scoped to city/county
  section the locality appears under.
"""

from __future__ import annotations

import argparse
import re
import sys
import urllib.error
import urllib.request
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "va"
STATE = "VA"
BUCKET = "va"
SOURCE_PORTAL_URL = "https://www.apa.virginia.gov/local-government/reports?type=comparative-reports"
PUBLISHER = "Virginia Auditor of Public Accounts"
DOCUMENT_TYPE = "apa_comparative_report_xlsx"
TOPLINE_DEFINITION = (
    "APA Comparative Report of Local Government Revenues & Expenditures, "
    "Exhibit C col 22 (Education / Exhibit C-6) — total education "
    "expenditures per locality (Instruction + Admin + Pupil Transport + "
    "O&M + other education functions)"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

# APA file URLs use opaque GUIDs that change per FY release. Hard-code known
# good URLs; the catalog page must be re-scraped manually for new years.
KNOWN_FILE_URLS: dict[int, str] = {
    2025: "https://dlasprodpublic.blob.core.windows.net/apa/093D6F15-1079-4D72-87B5-CBBFEF1FDA47.xlsx",
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=120) as resp:
        return resp.read()


# Phase 7.5 — VA Exhibit C6 column mapping. Exhibit C6 has rich
# functional breakdown of the Education topline (col 22 of Exhibit C):
# 5 expenditure categories + 3 revenue categories = 8 canonical
# categories per locality.
VA_C6_COLS: dict[str, tuple[list[int], str]] = {
    "instruction": ([3], "APA Exhibit C6 'Instruction' column"),
    "administration": (
        [7],
        "APA Exhibit C6 'Administration, Attendance and Health' column "
        "(includes health services; VA bundles these together)",
    ),
    "transportation": ([11], "APA Exhibit C6 'Pupil Transportation Services' column"),
    "operations_maintenance": ([15], "APA Exhibit C6 'Operation and Maintenance Services' column"),
    "food_service": (
        [19],
        "APA Exhibit C6 'School Food Services and Other Non-Instructional Operations' column",
    ),
    "revenue_state": ([28], "APA Exhibit C6 'Commonwealth Categorical Aid' column"),
    "revenue_federal": (
        [30, 32],
        "APA Exhibit C6 'Federal Pass-Through' + 'Direct Federal Aid' columns",
    ),
    "revenue_local": ([34], "APA Exhibit C6 'Local Charges for Service' column"),
}


def parse_c6(xlsx_bytes: bytes) -> dict[tuple[str, str], dict[str, float]]:
    """Parse Exhibit C6 — per-(section, locality) component amounts.

    Same section-tracking logic as Exhibit C; Exhibit C6 has the locality
    name in col 2 and category amounts in cols 3, 7, 11, 15, 19, 28, 30,
    32, 34. Two header rows precede the data; section headers say
    'Locality City of:' / 'County of:' / 'Town of:'.
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "Exhibit C6" not in wb.sheetnames:
        return {}
    ws = wb["Exhibit C6"]
    out: dict[tuple[str, str], dict[str, float]] = {}
    section: str | None = None
    seen_grand_total = False
    for r in ws.iter_rows(values_only=True):
        if not r:
            continue
        c2 = r[2] if len(r) > 2 else None
        if not c2:
            continue
        cell = str(c2)
        if "Locality" in cell:
            if "City of" in cell:
                section = "city"
            elif "County of" in cell:
                section = "county"
            elif "Town of" in cell:
                section = "town"
            continue
        if cell in ("Total", "Grand Total"):
            if cell == "Grand Total":
                seen_grand_total = True
            continue
        if seen_grand_total:
            break
        if not isinstance(r[0], (int, float)):
            continue
        if section is None:
            continue
        name = re.sub(r"[*#††]+$", "", cell.strip()).strip()
        components: dict[str, float] = {}
        for category, (col_idxs, _def) in VA_C6_COLS.items():
            v_sum = 0.0
            for ci in col_idxs:
                v = r[ci] if ci < len(r) else None
                if v is None:
                    continue
                try:
                    v_sum += float(v)
                except (TypeError, ValueError):
                    pass
            if v_sum > 0:
                components[category] = v_sum
        if components:
            out[(section, name.upper())] = components
    return out


def parse_education_expenditures(xlsx_bytes: bytes) -> list[dict]:
    """Parse Exhibit C: walk rows tracking section ('city' | 'county' | 'town'),
    extract (section, locality_name, education_total) for each numbered row."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb["Exhibit C"]
    rows = list(ws.iter_rows(values_only=True))

    # Education column is col 22 (0-indexed)
    EDUCATION_COL = 22

    out: list[dict] = []
    section: str | None = None
    seen_grand_total = False
    for r in rows:
        if not r or not r[1]:
            continue
        cell1 = str(r[1])
        # Section header: "Locality \nCity of:" / "Locality \nCounty of:" / "Locality \nTown of:"
        if "Locality" in cell1:
            if "City of" in cell1:
                section = "city"
            elif "County of" in cell1:
                section = "county"
            elif "Town of" in cell1:
                section = "town"
            continue
        # Skip Total / Grand Total summary rows
        if cell1 in ("Total", "Grand Total"):
            if cell1 == "Grand Total":
                seen_grand_total = True
            continue
        if seen_grand_total:
            break
        # Skip rows where col 0 isn't a number (header artifacts)
        if not isinstance(r[0], (int, float)):
            continue
        if section is None:
            continue
        edu = r[EDUCATION_COL]
        if edu is None:
            continue
        try:
            edu_val = float(edu)
        except (TypeError, ValueError):
            continue
        if edu_val <= 0:
            continue
        # Locality name: strip footnote chars (* # etc.) and whitespace
        name = re.sub(r"[*#††]+$", "", cell1.strip()).strip()
        out.append({
            "section": section,
            "name": name,
            "education_total": edu_val,
        })
    return out


def build_va_crosswalk(client: Client) -> dict[tuple[str, str], dict]:
    """{(section, name_upper): district_row}. The same locality name
    can appear once as a city and once as a county (e.g. 'Fairfax' city +
    'Fairfax' county) — section disambiguates."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[tuple[str, str], dict] = {}
    for r in rows:
        name = (r.get("lea_name") or "").upper()
        # VDOE division names follow the pattern:
        #   "{Locality} City Public Schools" for cities (e.g. "Alexandria City Public Schools")
        #   "{Locality} County Public Schools" for counties (e.g. "Fairfax County Public Schools")
        m_city = re.match(r"^(.+?)\s+CITY\s+PUBLIC SCHOOLS$", name)
        m_county = re.match(r"^(.+?)\s+COUNTY\s+PUBLIC SCHOOLS$", name)
        if m_city:
            out[("city", m_city.group(1).strip())] = r
        elif m_county:
            out[("county", m_county.group(1).strip())] = r
        else:
            # Some divisions don't follow the pattern (e.g. "Colonial Beach", "West Point")
            # — index them by raw name under both sections.
            out[("city", name)] = r
            out[("county", name)] = r
    return out


def extract(*, fiscal_year: int = 2025, triggered_by: str = "manual") -> dict:
    print(f"VA extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No APA Comparative Report URL for fiscal_year={fiscal_year}; "
                "add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/',1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.1f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/apa_comparative_report.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'Exhibit C'; col 22 'Education (Exhibit C-6)' per "
                "locality; section header (City of: / County of:) "
                "disambiguates same-name city vs county"
            ),
            notes=f"FY{fiscal_year} APA Comparative Report (year ended June 30, {fiscal_year})",
        )

        crosswalk = build_va_crosswalk(client)
        print(f"  VA crosswalk: {len(crosswalk):,} (section, name) keys "
              f"covering {len({d['leaid'] for d in crosswalk.values()})} unique LEAs")

        rows_data = parse_education_expenditures(xlsx_bytes)
        print(f"  APA Exhibit C localities (with Education > 0): {len(rows_data):,}")

        # Phase 7.5 — parse Exhibit C6 component breakdown
        c6_components = parse_c6(xlsx_bytes)
        print(f"  APA Exhibit C6 localities: {len(c6_components):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for row in rows_data:
            key = (row["section"], row["name"].upper())
            district = crosswalk.get(key)
            if district is None:
                no_match.append(f"{row['section']}:{row['name']}")
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=row["education_total"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            # Phase 7.5 — emit components from Exhibit C6
            comp_amts = c6_components.get(key, {})
            components: list[ComponentInput] = []
            for category, amount in comp_amts.items():
                col_idxs, definition = VA_C6_COLS[category]
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=definition,
                        line_or_cell_reference=(
                            f"Sheet 'Exhibit C6'; cols {col_idxs} on row for "
                            f"{row['section']}/{row['name']}"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched localities: {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )
        if no_match[:10]:
            print(f"  sample unmatched: {no_match[:10]}")

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2025)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
