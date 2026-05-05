"""California Budget extractor — CDE SACS Data Viewer (Budget July 1).

Source: https://viewer.sacs-cde.org/  (Angular SPA backed by a JSON API)

This is the BS1 ("Budget, July 1") submission cycle. Districts adopt the
budget by July 1 (Cal. Educ. Code § 42127) and submit to the SACS Data
Viewer. Companion to extractors/ca.py which pulls the Unaudited Actuals
(reportingPeriod="A") for completed years.

API discovered by reading the SPA's main.js bundle:

  GET  /api/ReferenceData/ActiveFiscalYears
       → list of fiscal years with active reporting periods
  POST /api/Entities/Items
       body: {"request":{"data":{"caFiscalYear":N,"entityType":"SchoolDistrict"},
              "first":0,"rows":<page>,"sorts":[],"filterFields":[],"filters":[],
              "globalFilter":null},"runMode":null,"testRunId":null,
              "timeZoneId":"America/Los_Angeles"}
       → list of LEAs with 14-digit cdsCode (CountyCode2 + DistrictCode5 + SchoolCode7)
  POST /api/SubmissionArtifacts/Items
       body.data: {"fullFiscalYear":"YYYY-YY","reportingPeriod":"BS1",
                   "cdsCode":"<14 digits>","excludeArtifactTypes":[]}
       → list of artifacts. We want type='Data' (single XLSX per submission).
  GET  /api/SubmissionArtifact/{id}/Blob
       → the binary file (XLSX/DAT/PDF/ZIP)

Topline definition (matches extractors/ca.py for actuals comparability):
  Object 1000-7999 in Funds 01-29, ColumnCode='BB' (this-year Budget),
  FullFiscalYear=target FY label.

Fiscal year conversion:
  our schema fiscal_year=N (= ending year of school year)
    ↔ caFiscalYear=N-1 (CA's "starting year")
    ↔ fullFiscalYear="<N-1 mod 100>-<N mod 100>"
  e.g. fiscal_year=2026 → caFiscalYear=2025, fullFiscalYear="2025-26"
"""

from __future__ import annotations

import argparse
import json
import sys
import time
import urllib.error
import urllib.request
from io import BytesIO
from typing import Any

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "ca_budget"
STATE = "CA"
BUCKET = "ca"
VIEWER_BASE = "https://viewer.sacs-cde.org"
PUBLISHER = "California Department of Education"
DOCUMENT_TYPE = "sacs_budget_july1_xlsx"
TOPLINE_DEFINITION = (
    "CDE SACS Budget July 1 (BS1), UserGL ColumnCode='BB' Object 1000-7999 "
    "in Funds 01-29 (governmental funds), summed per LEA"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"


def _fy_codes(fiscal_year: int) -> tuple[int, str]:
    """fiscal_year=2026 → (caFiscalYear=2025, fullFiscalYear='2025-26')."""
    ca_fy = fiscal_year - 1
    full = f"{ca_fy:04d}-{fiscal_year % 100:02d}"
    return ca_fy, full


def _post_json(path: str, body: dict[str, Any], timeout: float = 30) -> dict[str, Any]:
    req = urllib.request.Request(
        f"{VIEWER_BASE}{path}",
        data=json.dumps(body).encode(),
        headers={
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": USER_AGENT,
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode())


def _get_blob(path: str, timeout: float = 60) -> bytes:
    req = urllib.request.Request(
        f"{VIEWER_BASE}{path}", headers={"User-Agent": USER_AGENT}
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def _envelope(data: dict[str, Any], rows: int = 1000) -> dict[str, Any]:
    """Wrap a `vC`-shape data dict in the FilterSortPageRequest+q0 envelope
    the SPA uses (see main.js: q0.create + Sn.createFilterSortPageRequest)."""
    return {
        "request": {
            "data": data,
            "first": 0,
            "rows": rows,
            "sorts": [],
            "filterFields": [],
            "filters": [],
            "globalFilter": None,
        },
        "runMode": None,
        "testRunId": None,
        "timeZoneId": "America/Los_Angeles",
    }


def fetch_entities(
    ca_fiscal_year: int, entity_types: tuple[str, ...] = ("SchoolDistrict", "CharterSchool")
) -> list[dict[str, Any]]:
    """Return SACS entities tagged with their entityType. Each cdsCode is
    14-digit:
      - SchoolDistrict: CountyCode(2) + DistrictCode(5) + 0000000
        — match by cdsCode[:7] against master state_leaid suffix.
      - CharterSchool:  CountyCode(2) + AuthorizerDistrictCode(5) + SchoolCode(7)
        — match by cdsCode[7:14] (the 7-digit SchoolCode is what NCES uses
        as state_leaid for charter LEAs).
    """
    out: list[dict[str, Any]] = []
    for et in entity_types:
        body = _envelope(
            {"caFiscalYear": ca_fiscal_year, "entityType": et}, rows=2000
        )
        resp = _post_json("/api/Entities/Items", body)
        if resp.get("status") != "Success":
            raise RuntimeError(f"Entities/Items failed for {et}: {resp}")
        for r in resp["response"]["results"]:
            r["entityType"] = et
            out.append(r)
    return out


def find_data_artifact(
    full_fy: str, cds_code: str
) -> tuple[str | None, str | None]:
    """Return (artifact_id, submissionNumber) for the Data Extract XLSX of the
    BS1 submission, or (None, None) if not found."""
    body = _envelope(
        {
            "fullFiscalYear": full_fy,
            "reportingPeriod": "BS1",
            "cdsCode": cds_code,
            "excludeArtifactTypes": [],
        },
        rows=1000,
    )
    resp = _post_json("/api/SubmissionArtifacts/Items", body)
    if resp.get("status") != "Success":
        return None, None
    for r in resp["response"]["results"]:
        if r.get("type") == "Data" and r.get("fileFormat") == "XLSX":
            return r["id"], r.get("submissionNumber")
    return None, None


def parse_data_extract_topline(
    xlsx_bytes: bytes, full_fy: str
) -> float | None:
    """Sum UserGL Amount where ColumnCode='BB', FundCode 01-29, Object 1000-7999
    for FullFiscalYear == full_fy (the budget year)."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "UserGL" not in wb.sheetnames:
        return None
    ws = wb["UserGL"]
    total = 0.0
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # header
    for row in rows:
        if not row or len(row) < 13:
            continue
        fy_val, col_code = row[2], row[4]
        if fy_val != full_fy or col_code != "BB":
            continue
        fund_str, obj_str, amount = row[6], row[11], row[12]
        try:
            fund = int(fund_str)
            obj = int(obj_str)
        except (TypeError, ValueError):
            continue
        if 1 <= fund <= 29 and 1000 <= obj <= 7999:
            try:
                total += float(amount or 0)
            except (TypeError, ValueError):
                continue
    return total if total else None


def build_ca_crosswalk(client: Client) -> dict[str, dict]:
    """7-digit suffix of state_leaid → district row.

    Note: for CA SchoolDistrict LEAs the 7-digit suffix is the County+District
    portion of the CDS code; for CA CharterSchool LEAs that file as their own
    LEA, NCES uses the 7-digit SchoolCode portion as state_leaid. Both are
    unique within CA, so a single dict-by-suffix works for both entity types
    — we just need to extract the right substring from the SACS cdsCode at
    match time (see CDS_KEY_BY_TYPE below)."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("CA-"):
            out[sl.removeprefix("CA-")] = r
    return out


def cds_lookup_key(cds_code: str, entity_type: str) -> str:
    """Map a 14-digit SACS cdsCode + entityType to the 7-digit master state_leaid suffix."""
    if entity_type == "SchoolDistrict":
        return cds_code[:7]
    if entity_type == "CharterSchool":
        return cds_code[7:14]
    return cds_code[:7]  # fall back to district behavior for unknown types


def extract(*, fiscal_year: int = 2026, triggered_by: str = "manual",
            limit: int | None = None, sleep_between: float = 0.2) -> dict:
    ca_fy, full_fy = _fy_codes(fiscal_year)
    print(f"CA Budget extract: fiscal_year={fiscal_year} (caFY={ca_fy}, fullFY={full_fy})")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        crosswalk = build_ca_crosswalk(client)
        print(f"  crosswalk: {len(crosswalk):,} CA operating LEAs in master")

        entities = fetch_entities(ca_fy)
        sd_count = sum(1 for e in entities if e["entityType"] == "SchoolDistrict")
        ch_count = sum(1 for e in entities if e["entityType"] == "CharterSchool")
        print(f"  SACS entities for caFY {ca_fy}: "
              f"{sd_count:,} SchoolDistrict + {ch_count:,} CharterSchool = {len(entities):,}")

        no_match: list[str] = []
        no_artifact: list[str] = []
        no_topline: list[str] = []
        api_errors: list[str] = []
        processed = 0

        candidates: list[tuple[dict[str, Any], str, dict]] = []
        for ent in entities:
            full_cds = ent["cdsCode"]
            key = cds_lookup_key(full_cds, ent["entityType"])
            district = crosswalk.get(key)
            if district is not None:
                candidates.append((ent, key, district))
            else:
                no_match.append(f"{ent['entityType']}:{full_cds}")

        print(
            f"  matched {len(candidates):,} of {len(entities):,} SACS entities "
            f"to {len(crosswalk):,} master LEAs ({len(no_match):,} unmatched)"
        )
        if limit is not None:
            candidates = candidates[:limit]
            print(f"  --limit applied: capping to first {limit} candidates")

        for i, (ent, match_key, district) in enumerate(candidates, start=1):
            full_cds = ent["cdsCode"]
            label = f"[{i}/{len(candidates)}] {ent['name']} ({ent['entityType']} {full_cds})"
            try:
                art_id, sub_num = find_data_artifact(full_fy, full_cds)
            except Exception as e:
                api_errors.append(f"{match_key}: {type(e).__name__}: {e}")
                print(f"  {label}  API error on Items: {e}")
                continue

            if art_id is None:
                no_artifact.append(match_key)
                continue

            try:
                xlsx = _get_blob(f"/api/SubmissionArtifact/{art_id}/Blob")
            except Exception as e:
                api_errors.append(f"{match_key}: blob {type(e).__name__}: {e}")
                print(f"  {label}  blob error: {e}")
                continue

            topline = parse_data_extract_topline(xlsx, full_fy)
            if topline is None:
                no_topline.append(match_key)
                continue

            content_hash = sha256_bytes(xlsx)
            storage_relpath = (
                f"fy{fiscal_year}/budget/{full_cds}_{sub_num or art_id[:8]}.xlsx"
            )

            existing_src = (
                client.table("source_documents")
                .select("id")
                .eq("content_hash_sha256", content_hash)
                .execute()
            )
            if not existing_src.data:
                upload_source_document(
                    client=client,
                    bucket=BUCKET,
                    storage_path=storage_relpath,
                    content=xlsx,
                    mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            src_id = upsert_source_document_row(
                client=client,
                content_hash=content_hash,
                source_url=f"{VIEWER_BASE}/api/SubmissionArtifact/{art_id}/Blob",
                storage_path=f"{BUCKET}/{storage_relpath}",
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                publisher=PUBLISHER,
                document_type=DOCUMENT_TYPE,
                line_or_cell_reference=(
                    "Sheet 'UserGL'; sum Amount where FullFiscalYear="
                    f"'{full_fy}' AND ColumnCode='BB' AND FundCode in 01-29 "
                    "AND ObjectCode in 1000-7999"
                ),
                notes=(
                    f"FY{fiscal_year} adopted budget (BS1 reporting period). "
                    f"SACS submissionNumber={sub_num}, artifact_id={art_id}"
                ),
            )

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="adopted",
                topline_amount=topline,
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            _, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1
            processed += 1
            if processed % 25 == 0:
                print(f"  {label}  ${topline:,.0f} (changed={changed})")

            if sleep_between:
                time.sleep(sleep_between)

        print(
            f"  done. inserted/changed={run.records_changed} extracted={run.records_extracted} "
            f"no-artifact={len(no_artifact)} no-topline={len(no_topline)} "
            f"unmatched-entities={len(no_match)} api-errors={len(api_errors)}"
        )
        if api_errors[:5]:
            print(f"  api errors sample: {api_errors[:3]}")

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_artifact": no_artifact,
        "no_topline": no_topline,
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2026)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    p.add_argument("--limit", type=int, default=None,
                   help="cap districts processed (for smoke testing)")
    p.add_argument("--sleep", type=float, default=0.2,
                   help="seconds to sleep between districts")
    args = p.parse_args()
    extract(
        fiscal_year=args.fiscal_year,
        triggered_by=args.triggered_by,
        limit=args.limit,
        sleep_between=args.sleep,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
