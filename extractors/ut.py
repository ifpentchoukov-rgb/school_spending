"""Utah extractor — USBE Annual Financial Report Summary Expenditure (AF).

Source: https://www.schools.utah.gov/financialoperations/reporting.php
File: AFR Summary Expenditure AF.xlsx — published annually by the Utah
State Board of Education after AFR reconciliation. Latest as of
2026-05-05 is FY24 (file at .../2024fiscalyear/fy24annualfinancialreportafr/).

What this gives us:
  - Per-LEA all-funds-governmental Grand Total expenditure for completed
    Utah FYs. Sheets per fund (10 = General, 20 = LEA Foundation, 21 =
    Student Activity, etc.) plus 'Gov Funds Total' (the aggregated
    Governmental Funds Total) and 'Proprietary Funds Total'.
  - 41 districts + 115 charter LEAs = 156 LEAs in the FY24 file.

Topline definition:
  'Gov Funds Total' sheet, col 38 'Grand Total' — sum of all governmental
  funds operating expenditure across the 5 functional categories
  (Instruction, Support Services, Noninstructional Services, Facilities
  Acquisition & Construction, Debt Service & Misc) and 8 object
  subcategories. Aligned with F-33 'total expenditures' frame.

Status: `actual` — post-AFR audited.

Crosswalk (v1: districts only):
  Master state_leaid format: 'UT-{NN}' (2-digit zero-padded) for districts
                              'UT-A{N}' for charter LEAs (3 in master).
  AFR file:                  LeaType + LeaNbr (e.g. District 1 = Alpine).
  → For LeaType='District', key = f"{LeaNbr:02d}".
  Charter coverage TBD; needs LeaNbr → master charter A-code map.
"""

from __future__ import annotations

import argparse
import sys
import urllib.error
import urllib.request
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "ut"
STATE = "UT"
BUCKET = "ut"
SOURCE_PORTAL_URL = "https://www.schools.utah.gov/financialoperations/reporting.php"
PUBLISHER = "Utah State Board of Education"
DOCUMENT_TYPE = "usbe_afr_summary_expenditure_xlsx"
TOPLINE_DEFINITION = (
    "USBE AFR Summary Expenditure 'Gov Funds Total' sheet, col 38 "
    "'Grand Total' — all-funds governmental operating expenditure per LEA"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

# AFR file paths vary slightly by FY. Hard-code known good URLs.
KNOWN_FILE_URLS: dict[int, str] = {
    2024: "https://www.schools.utah.gov/financialoperations/reporting/reports/annualfinancialreport/2024fiscalyear/fy24annualfinancialreportafr/AFR%20Summary%20Expenditure%20AF.xlsx",
    2023: "https://www.schools.utah.gov/financialoperations/reporting/reports/annualfinancialreport/2023fiscalyear/AFR%20Summary%20Expenditure.xlsx",
}

GRAND_TOTAL_COL_IDX = 38  # col 38 in Gov Funds Total sheet
LEATYPE_COL_IDX = 0
LEANBR_COL_IDX = 1
LEA_COL_IDX = 2

# Phase 7.5 — UT Function-category 0-indexed col ranges (data row).
# Row 4 of the sheet labels these spans: 1xxx Instruction (cols 3-10, 8 obj
# cols), 2xxx Support Services (11-18), 3xxx Noninstructional (19-26),
# 4xxx Facilities Acquisition (27-34, capital_outlay), 5xxx Debt Service
# & Misc (35-37, 3 obj cols only).
_UT_FUNCTION_BLOCKS: list[tuple[str, range]] = [
    ("instruction", range(3, 11)),
    ("support_services_instruction", range(11, 19)),
    ("food_service", range(19, 27)),  # Noninstructional ≈ food + community
    ("capital_outlay", range(27, 35)),
    ("debt_service", range(35, 38)),
]
# Object 2xx (Employee Benefits) lives at offset +1 within each 8-col block.
_UT_BENEFITS_COLS: tuple[int, ...] = (4, 12, 20, 28)


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def parse_gov_funds_total(xlsx_bytes: bytes) -> list[dict]:
    """Return [{lea_type, lea_nbr, lea_name, total}, ...] from Gov Funds Total.

    Uses non-readonly mode because the AFR sheet has merged cells in the
    header rows that read_only mode doesn't expose consistently."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    if "Gov Funds Total" not in wb.sheetnames:
        raise RuntimeError(f"'Gov Funds Total' sheet missing; sheets={wb.sheetnames}")
    ws = wb["Gov Funds Total"]
    # Find the Grand Total column by scanning row 4 (function-category header)
    # in the non-readonly worksheet.
    grand_total_col = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=col).value
        if v == "Grand Total":
            grand_total_col = col - 1  # convert to 0-indexed for tuples below
            break
    if grand_total_col is None:
        raise RuntimeError(
            f"'Grand Total' column not found in row 4; max_col={ws.max_column}"
        )

    out: list[dict] = []
    # Data starts at row 6 (1-indexed); rows are tuples via iter_rows
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or not row[LEATYPE_COL_IDX] or row[LEANBR_COL_IDX] is None:
            continue
        try:
            total = float(row[grand_total_col]) if row[grand_total_col] is not None else None
        except (TypeError, ValueError):
            continue
        if not total or total <= 0:
            continue
        try:
            lea_nbr = int(float(row[LEANBR_COL_IDX]))
        except (TypeError, ValueError):
            continue

        def _sum_cols(cols):
            s = 0.0
            for c in cols:
                if c >= len(row):
                    continue
                v = row[c]
                if v is None:
                    continue
                try:
                    s += float(v)
                except (TypeError, ValueError):
                    pass
            return s

        components: dict[str, float] = {}
        for category, col_range in _UT_FUNCTION_BLOCKS:
            v = _sum_cols(col_range)
            if v > 0:
                components[category] = v
        # employee_benefits = sum of Object 2xx across all function blocks
        benefits = _sum_cols(_UT_BENEFITS_COLS)
        if benefits > 0:
            components["employee_benefits"] = benefits

        out.append({
            "lea_type": str(row[LEATYPE_COL_IDX]).strip(),
            "lea_nbr": lea_nbr,
            "lea_name": row[LEA_COL_IDX],
            "total": total,
            "components": components,
        })
    return out


def build_ut_crosswalk(client: Client) -> dict[str, dict]:
    """{state_leaid_suffix: district_row}. Suffixes are '01'..'99' for
    districts and 'A1'..'AN' for charter LEAs."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("UT-"):
            out[sl.removeprefix("UT-")] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"UT extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No USBE AFR URL for fiscal_year={fiscal_year}; add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/',1)[-1].split('?')[0]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e3:.1f} KB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/afr_summary_expenditure.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'Gov Funds Total'; LeaType='District' filter; "
                "key = f\"{LeaNbr:02d}\" matches state_leaid suffix; "
                "topline = col 38 'Grand Total'"
            ),
            notes=(
                f"FY{fiscal_year} USBE AFR Summary Expenditure (audited "
                "all-funds governmental operating expenditure). Charters "
                "in this file (LeaType='Charter') are not yet matched to "
                "master UT-A* codes — sibling crosswalk TBD."
            ),
        )

        crosswalk = build_ut_crosswalk(client)
        print(f"  UT crosswalk: {len(crosswalk):,} state→NCES mappings")

        afr_rows = parse_gov_funds_total(xlsx_bytes)
        n_districts = sum(1 for r in afr_rows if r["lea_type"] == "District")
        n_charters = sum(1 for r in afr_rows if r["lea_type"] == "Charter")
        print(f"  AFR LEAs: {len(afr_rows)} ({n_districts} District + {n_charters} Charter)")

        no_match: list[str] = []
        skipped_charter = 0
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for row in afr_rows:
            if row["lea_type"] != "District":
                # v1 handles districts only; charters need separate crosswalk
                skipped_charter += 1
                continue
            key = f"{row['lea_nbr']:02d}"
            district = crosswalk.get(key)
            if district is None:
                no_match.append(f"{row['lea_type']} {row['lea_nbr']} {row['lea_name']}")
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=row["total"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            components: list[ComponentInput] = []
            for category, amount in row.get("components", {}).items():
                if amount <= 0:
                    continue
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=(
                            f"USBE AFR Summary Expenditure 'Gov Funds Total' "
                            f"sheet, LEA {row['lea_nbr']:02d}: sum of object "
                            f"cols within function block mapping to '{category}' "
                            f"(or Object 2xx across blocks for employee_benefits)"
                        ),
                        line_or_cell_reference=(
                            f"Sheet 'Gov Funds Total'; LeaNbr={row['lea_nbr']}; "
                            f"per _UT_FUNCTION_BLOCKS / _UT_BENEFITS_COLS"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"skipped charters: {skipped_charter}; unmatched districts: {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024,
                   help="USBE AFR FY (latest as of 2026-05-05: 2024)")
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
