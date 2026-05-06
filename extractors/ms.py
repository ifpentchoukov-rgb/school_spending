"""Mississippi extractor — MDE Superintendent's Annual Report by Functional Area XLSX.

Source: https://mdek12.org/mbe/superintendent2024/
File: 2023-2024-Expenditure-Totals-for-Public-Schools-by-Functional-Area_FINAL.xlsx
      Published with the Superintendent's Annual Report each fall.

What this gives us:
  - Per-district expenditure breakdown across 5 functional areas
    (Instruction, General Administration, School Administration, Other
    Instructional Support, Other Non-Instructional) plus a pre-summed
    'Total Current Operational Expenses' column. ~196 rows covering
    137 traditional districts + ~58 charter LEAs and special schools.

Topline definition:
  Column 'Total Current Operational Expenses' (col 19, 0-indexed) per
  district. Excludes Capital (col 20 'Capitalized Equipment
  Expenditures') and debt service. Aligned with F-33 'current
  expenditures' frame.

Status: `actual` — post-AFR audited.

Crosswalk:
  Master state_leaid format: 'MS-{4-digit}' (e.g. 'MS-0200' Alcorn)
  XLSX 'Dist No':            integer (e.g. 200)
  → state_leaid suffix == zfill(Dist No, 4).

Note: MDE Azure Application Gateway requires a real browser-style
Referer header for these assets; bare downloads return 403.
"""

from __future__ import annotations

import argparse
import io
import sys
import urllib.error
import urllib.request

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "ms"
STATE = "MS"
BUCKET = "ms"
SOURCE_PORTAL_URL = "https://mdek12.org/mbe/superintendent2024/"
PUBLISHER = "Mississippi Department of Education (Office of School Financial Services)"
DOCUMENT_TYPE = "mde_supt_annual_report_func_area_xlsx"
TOPLINE_DEFINITION = (
    "MDE Superintendent's Annual Report 'Expenditure Totals for "
    "Public Schools by Functional Area' XLSX, column 'Total Current "
    "Operational Expenses' per district. Excludes Capitalized "
    "Equipment Expenditures and debt service. Aligned with F-33 "
    "'current expenditures' frame."
)
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

KNOWN_FILE_URLS: dict[int, str] = {
    # FY24 = SY 2023-24; published Nov 2024.
    2024: "https://mdek12.org/wp-content/uploads/sites/29/2024/11/2023-2024-Expenditure-Totals-for-Public-Schools-by-Functional-Area_FINAL.xlsx",
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Referer": SOURCE_PORTAL_URL,
        },
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def parse_ms(xlsx_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    # Sheet name has a trailing space in the published file
    sheet = wb.sheetnames[0]
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    # Validate the crucial column at index 19.
    if not (header and len(header) > 19 and isinstance(header[19], str)
            and "Total Current Operational" in header[19]):
        raise RuntimeError(
            f"Unexpected MS XLSX layout; col 19 was {header[19] if len(header)>19 else None!r}"
        )
    out: list[dict] = []
    for r in rows:
        if not r or r[0] is None:
            continue
        try:
            d = int(r[0])
        except (TypeError, ValueError):
            continue
        amt = r[19]
        if amt is None:
            continue
        try:
            v = float(amt)
        except (TypeError, ValueError):
            continue
        if v <= 0:
            continue
        out.append({"code": f"{d:04d}", "total_op_exp": v})
    return out


def build_ms_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("MS-"):
            out[sl.removeprefix("MS-")] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"MS extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No MDE Sup Annual Report URL for fiscal_year={fiscal_year}; "
                "add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/', 1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.2f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/expenditure_by_functional_area.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "First (only) sheet; col 19 'Total Current Operational "
                "Expenses' per row; match zfill(Dist No, 4) == "
                "state_leaid suffix"
            ),
            notes=f"FY{fiscal_year} MDE Sup Annual Report - Expenditures by Functional Area",
        )

        crosswalk = build_ms_crosswalk(client)
        print(f"  MS crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_data = parse_ms(xlsx_bytes)
        print(f"  MDE districts with FY{fiscal_year} data: {len(district_data):,}")

        no_match: list[str] = []
        for d in district_data:
            district = crosswalk.get(d["code"])
            if district is None:
                no_match.append(d["code"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            _, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched MDE codes (charters/special schools): {len(no_match)}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
