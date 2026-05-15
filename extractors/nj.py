"""New Jersey extractor — NJDOE Taxpayers' Guide to Education Spending (TGES).

Source: https://www.nj.gov/education/guide/
File: Detail_FY{NN}.xlsx (one per FY) — published in the year after the
FY ends. Latest as of 2026-05-05 is Detail_FY24.xlsx (covers SY 2023-24).

What this gives us:
  - Per-district per-pupil and total spending breakdown for ~671 NJ
    educational entities (regular school districts + vocational +
    county special services + charter LEAs + state-operated +
    educational services commissions).
  - NJDOE compiles from each district's CAFR (Comprehensive Annual
    Financial Report).

Topline definition:
  'Total Spending' column from the 'Detail FY{NN}' sheet — total
  district expenditure including general current expense, capital
  outlay, grants & entitlements, food services, and debt service.
  Aligned with F-33 'total expenditures' frame.

Status: `actual` — post-CAFR audited numbers.

Crosswalk:
  Master state_leaid format: 'NJ-{2-digit-County}{4-digit-District}'
                             (e.g. 'NJ-010110' Atlantic City)
  TGES file:                 'County' (text name) + 'District Code' (3-4 digit)
  → Look up county code (NJ_COUNTY_CODES); zero-pad district code to 4
    digits; concatenate. Atlantic + 110 → 01 + 0110 → 010110.
"""

from __future__ import annotations

import argparse
import sys
import urllib.error
import urllib.request
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "nj"
STATE = "NJ"
BUCKET = "nj"
SOURCE_PORTAL_URL = "https://www.nj.gov/education/guide/"
PUBLISHER = "New Jersey Department of Education"
DOCUMENT_TYPE = "njdoe_tges_detail_xlsx"
TOPLINE_DEFINITION = (
    "NJDOE TGES Detail FY{NN} sheet, 'Total Spending' column — total "
    "district expenditure (general current expense + capital outlay + "
    "grants/entitlements + food services + debt service)"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

# NJDOE 2-digit county codes (alphabetical).
NJ_COUNTY_CODES: dict[str, str] = {
    "Atlantic": "01", "Bergen": "03", "Burlington": "05", "Camden": "07",
    "Cape May": "09", "Cumberland": "11", "Essex": "13", "Gloucester": "15",
    "Hudson": "17", "Hunterdon": "19", "Mercer": "21", "Middlesex": "23",
    "Monmouth": "25", "Morris": "27", "Ocean": "29", "Passaic": "31",
    "Salem": "33", "Somerset": "35", "Sussex": "37", "Union": "39",
    "Warren": "41",
    # State / multi-county special categories don't always match a single
    # 2-digit county; left out so they cleanly fall into no_match.
}

# TGES file URL pattern. The file lives under the year-after-FY release dir
# (e.g. /2025/Detail_FY24.xlsx for FY24 spending released in 2025).
KNOWN_FILE_URLS: dict[int, str] = {
    2024: "https://www.nj.gov/education/guide/docs/2025/Detail_FY24.xlsx",
    2023: "https://www.nj.gov/education/guide/docs/2025/Detail_FY23.xlsx",
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


# Phase 7.5 — NJ TGES Detail columns. Most cols are per-pupil; we
# multiply by daily enrollment to recover total dollars. Maps to 3
# canonical categories (universal floor: debt + capital + food).
# instruction/admin/etc. are folded into 'General Current Expense Per
# Pupil' which combines instruction + support + admin into one bucket
# (NJDOE doesn't separate further at TGES granularity). To break out
# instruction etc. we'd need to read the underlying CAFRs — out of scope
# for Phase 7.5.
NJ_PERPUPIL_COL_HEADERS: dict[str, list[str]] = {
    # category -> list of TGES column header(s) to sum (per-pupil $).
    "capital_outlay": ["Total Capital Outlay Per Pupil"],
    "food_service": ["Total Food Services \nPer Pupil"],
    "debt_service": [
        "Debt Service on Locally Issued Bonds Per Pupil",
        "Debt Service On School Development Authority Bonds Per Pupil",
    ],
}


def parse_tges_detail(xlsx_bytes: bytes) -> list[dict]:
    """Return [{county, name, code, total_spending, enrollment, perpupil_components}]."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, r in enumerate(rows):
        if r and r[0] == "County":
            header_idx = i
            break
    if header_idx is None:
        raise RuntimeError("'County' header row not found in TGES Detail")
    header = rows[header_idx]
    try:
        county_col = header.index("County")
        name_col = header.index("District Name")
        code_col = header.index("District Code")
    except ValueError as e:
        raise RuntimeError(f"Header missing column: {e}")
    # Find Total Spending column (variations across years).
    total_col = None
    for i, h in enumerate(header):
        if h and "Total Spending" in str(h):
            total_col = i
            break
    if total_col is None:
        raise RuntimeError(f"'Total Spending' column not found; header={header}")
    # Find enrollment column (for per-pupil → total reconstruction).
    enroll_col = None
    for i, h in enumerate(header):
        if h and "Enrollment" in str(h):
            enroll_col = i
            break
    # Map per-pupil headers → column indices for component reconstruction.
    perpupil_cols: dict[str, list[int]] = {}
    for category, possible_headers in NJ_PERPUPIL_COL_HEADERS.items():
        col_idxs: list[int] = []
        for h_pattern in possible_headers:
            for i, h in enumerate(header):
                if h and str(h).strip() == h_pattern.strip():
                    col_idxs.append(i)
                    break
        if col_idxs:
            perpupil_cols[category] = col_idxs

    out: list[dict] = []
    for r in rows[header_idx + 1:]:
        if not r or not r[county_col] or r[code_col] is None:
            continue
        try:
            total = float(r[total_col]) if r[total_col] is not None else None
        except (TypeError, ValueError):
            continue
        enrollment = None
        if enroll_col is not None:
            try:
                enrollment = float(r[enroll_col]) if r[enroll_col] is not None else None
            except (TypeError, ValueError):
                pass
        # Build per-pupil component map: {category: per_pupil_sum}
        pp_components: dict[str, float] = {}
        for category, col_idxs in perpupil_cols.items():
            v_sum = 0.0
            for ci in col_idxs:
                v = r[ci] if ci < len(r) else None
                if v is None:
                    continue
                try:
                    v_sum += float(v)
                except (TypeError, ValueError):
                    pass
            if v_sum > 0:
                pp_components[category] = v_sum
        if not total or total <= 0:
            continue
        try:
            code = int(float(r[code_col]))
        except (TypeError, ValueError):
            continue
        out.append({
            "county": str(r[county_col]).strip(),
            "name": str(r[name_col]).strip() if r[name_col] else None,
            "code": code,
            "total_spending": total,
            "enrollment": enrollment,
            "perpupil_components": pp_components,
        })
    return out


def build_nj_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("NJ-"):
            out[sl.removeprefix("NJ-")] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"NJ extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No NJ TGES URL for fiscal_year={fiscal_year}; add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/',1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e3:.1f} KB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/tges_detail.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'Detail FY{NN}'; key = NJ_COUNTY_CODES[County] + "
                "zfill(District Code, 4); topline = 'Total Spending' column"
            ),
            notes=f"FY{fiscal_year} NJ TGES Total Spending Detail",
        )

        crosswalk = build_nj_crosswalk(client)
        print(f"  NJ crosswalk: {len(crosswalk):,} state→NCES mappings")

        tges_rows = parse_tges_detail(xlsx_bytes)
        print(f"  TGES districts (total > 0): {len(tges_rows):,}")

        no_match: list[str] = []
        unknown_county: set[str] = set()
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for row in tges_rows:
            cc = NJ_COUNTY_CODES.get(row["county"])
            if not cc:
                unknown_county.add(row["county"])
                no_match.append(f"<unknown-county>:{row['county']}/{row['code']}")
                continue
            key = f"{cc}{row['code']:04d}"
            district = crosswalk.get(key)
            if district is None:
                no_match.append(key)
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=row["total_spending"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            # Phase 7.5 — emit components from per-pupil × enrollment.
            enrollment = row.get("enrollment")
            if enrollment and enrollment > 0:
                components: list[ComponentInput] = []
                for category, per_pupil in row.get("perpupil_components", {}).items():
                    src_headers = NJ_PERPUPIL_COL_HEADERS[category]
                    components.append(
                        ComponentInput(
                            category=category,
                            amount=float(per_pupil) * float(enrollment),
                            definition=(
                                f"NJDOE TGES Detail sheet: sum of per-pupil column(s) "
                                f"{src_headers}, multiplied by 'Daily Enrollment Plus "
                                f"Sent Pupils' to reconstruct total dollars"
                            ),
                            line_or_cell_reference=(
                                f"Row for {row['county']}/{row['code']:04d} "
                                f"({row['name']}); columns {src_headers}"
                            ),
                        )
                    )
                if components:
                    ins, upd, unch = upsert_components(
                        client=client,
                        budget_event_id=event_id,
                        components=components,
                    )
                    n_components_inserted += ins
                    n_components_updated += upd
                    n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched: {len(no_match)} (incl. {len(unknown_county)} unknown counties: {sorted(unknown_county)})"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024,
                   help="TGES file FY (latest as of 2026-05-05: 2024 = SY 2023-24)")
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
