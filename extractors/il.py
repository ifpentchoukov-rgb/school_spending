"""Illinois extractor — ISBE Operating Expense Per Pupil (OEPP) / Per Capita
Tuition Charge (PCTC) bulk Excel.

Source: https://www.isbe.net/Pages/Operating-Expense-Per-Pupil.aspx
File pattern: https://www.isbe.net/_layouts/Download.aspx?SourceUrl=
              /Documents/FY{NN}-OEPP-PCTC.xlsx
e.g. FY24-OEPP-PCTC.xlsx covers SY 2023-24 = our fiscal_year=2024.

What this gives us:
  - Audited actual Total Operating Expenditures per IL district per fiscal
    year. ISBE publishes the file once per year, typically June, after AFR
    reconciliation. As of 2026-05-05 the latest file is FY24-OEPP-PCTC.xlsx
    (~864 rows including elementary K-8 / high-school splits and coop
    entities; matches 385 of our 397 IL operating LEAs).

Topline definition:
  "Total Operating Expenditures" column from the OEPP-PCTC sheet. ISBE
  computes this from the AFR's Operating Funds (Educational, Operations &
  Maintenance, Transportation, IMRF/Social Security, Working Cash, Tort)
  net of inter-fund transfers. This aligns with F-33's "current
  expenditures" concept — comparable to the actuals we extract for
  TX/CA/FL.

Status: `actual` — these are post-audit numbers.

What this does NOT give us:
  - Adopted budgets (Form 50-39). Districts file Form 50-39 with ISBE; the
    bulk public download is queued as a separate extractor.
  - The "Annual Statement of Affairs" (ASA) PDFs which districts publish
    locally — separate per-district scraping path.

Crosswalk:
  Master state_leaid format:  'IL-{Region:02}-{County:03}-{District:04}-{Type:02}'
                                e.g. 'IL-15-016-2990-25' (Chicago)
  ISBE OEPP RCDT No format:    11-digit concatenation, no hyphens
                                e.g. '15016299025'  (also Chicago)
  → strip 'IL-' and remove hyphens.
"""

from __future__ import annotations

import argparse
import sys
import urllib.request
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "il"
STATE = "IL"
BUCKET = "il"
SOURCE_PORTAL_URL = "https://www.isbe.net/Pages/Operating-Expense-Per-Pupil.aspx"
PUBLISHER = "Illinois State Board of Education"
DOCUMENT_TYPE = "isbe_oepp_pctc_xlsx"
TOPLINE_DEFINITION = (
    "ISBE OEPP-PCTC sheet, Total Operating Expenditures column "
    "(Educational + O&M + Transportation + IMRF/SS + Working Cash + Tort, "
    "net of inter-fund transfers — matches AFR audited operating expense)"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"


def _fy_short(fiscal_year: int) -> str:
    """2024 → '24'. ISBE filename uses 2-digit FY."""
    return f"{fiscal_year % 100:02d}"


def file_url(fiscal_year: int) -> str:
    return (
        "https://www.isbe.net/_layouts/Download.aspx?SourceUrl="
        f"/Documents/FY{_fy_short(fiscal_year)}-OEPP-PCTC.xlsx"
    )


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def normalize_il_state_leaid(state_leaid: str) -> str:
    """'IL-15-016-2990-25' → '15016299025'."""
    return state_leaid.removeprefix("IL-").replace("-", "")


def build_il_crosswalk(client: Client) -> dict[str, dict]:
    """11-digit RCDT (no hyphens) → district row."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("IL-"):
            out[normalize_il_state_leaid(sl)] = r
    return out


def parse_oepp(xlsx_bytes: bytes) -> list[dict]:
    """Return list of {rcdt, type, county, name, total_op_exp, oepp, ada}.
    Header row contains exactly the columns ISBE has used since FY20+.
    """
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    for r in rows[1:]:
        rcdt = r[0]
        if not rcdt:
            continue
        try:
            total_op_exp = float(r[4]) if r[4] is not None else None
        except (TypeError, ValueError):
            total_op_exp = None
        if total_op_exp is None:
            continue
        out.append({
            "rcdt": str(rcdt),
            "type": r[1],
            "county": r[2],
            "name": r[3],
            "total_op_exp": total_op_exp,
            "oepp": r[5],
            "ada": r[8],
        })
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"IL extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        print(f"  downloading {url.rsplit('/',1)[-1]}...")
        try:
            xlsx_bytes = download(url)
        except urllib.error.HTTPError as e:
            print(f"  FAILED: {e}")
            raise
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e3:.1f} KB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/oepp_pctc.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Single sheet; filter RCDT No == 11-digit state_leaid suffix; "
                "topline column 'Total Operating Expenditures'"
            ),
            notes=(
                f"FY{fiscal_year} ISBE OEPP-PCTC bulk file; covers ~860 IL "
                "entities (district + K-8/HS split partials)"
            ),
        )

        crosswalk = build_il_crosswalk(client)
        print(f"  IL crosswalk: {len(crosswalk):,} state→NCES mappings")

        oepp_rows = parse_oepp(xlsx_bytes)
        print(f"  OEPP rows: {len(oepp_rows):,}")

        no_match: list[str] = []
        for row in oepp_rows:
            district = crosswalk.get(row["rcdt"])
            if district is None:
                no_match.append(row["rcdt"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=row["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            _, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched OEPP RCDT codes (K-8/HS partials, coops, etc.): {len(no_match)}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024,
                   help="ISBE OEPP file FY (latest published as of 2026-05-05: 2024)")
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
