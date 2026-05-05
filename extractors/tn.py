"""Tennessee extractor — TDOE Annual Statistical Report Table 51.

Source: https://www.tn.gov/education/districts/federal-programs-and-oversight/
        data/department-reports/{YYYY}-annual-statistical-report.html
File: 2024-25_ASR_Excel.zip — published annually February of the following
calendar year. Contains 50+ table Excel files inside; we use Table 51
(per-district current expenditures + operating expenditures + enrollment).

What this gives us:
  - Per-district `Total Operating Expenditures` for every TN LEA. ASR is
    the audited summary published per Tenn. Code Ann. § 49-1-211.

Topline definition:
  Table 51 col 3 'TOTAL OPERATING EXPENDITURES' — total operating
  spending per district (instruction + student services + instructional
  staff + admin + plant O&M + other current). Aligned with F-33
  'current expenditures' frame.

Status: `actual` — post-AFR audited.

Crosswalk:
  Master state_leaid format: 'TN-{5-digit}' (e.g. 'TN-00010' Anderson)
  ASR Table 51 col 0:        3-digit district code (e.g. '010')
  → zero-pad ASR code to 5 digits.
"""

from __future__ import annotations

import argparse
import io
import sys
import urllib.error
import urllib.request
import zipfile

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "tn"
STATE = "TN"
BUCKET = "tn"
SOURCE_PORTAL_URL = "https://www.tn.gov/education/districts/federal-programs-and-oversight/data/department-reports/2025-annual-statistical-report.html"
PUBLISHER = "Tennessee Department of Education"
DOCUMENT_TYPE = "tdoe_asr_table51_xlsx"
TOPLINE_DEFINITION = (
    "TDOE Annual Statistical Report Table 51 col 3 'TOTAL OPERATING "
    "EXPENDITURES' per district (audited current expenditures)"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

KNOWN_FILE_URLS: dict[int, str] = {
    2025: "https://www.tn.gov/content/dam/tn/education/documents/asr/2024-25_ASR_Excel.zip",
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def parse_table51(zip_bytes: bytes) -> list[dict]:
    """Extract Table 51 from the ASR zip; return [{code, name, total_op_exp}]."""
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    target = next(
        (n for n in zf.namelist() if "TABLE 51" in n.upper() and n.endswith(".xlsx")),
        None,
    )
    if not target:
        raise RuntimeError(f"Table 51 not found in zip; names={zf.namelist()[:5]}")
    xlsx_bytes = zf.read(target)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    out: list[dict] = []
    for r in rows:
        if not r or not r[0]:
            continue
        code = str(r[0]).strip()
        # Skip non-data rows: code must be all digits (or contain digits)
        if not code or not any(ch.isdigit() for ch in code):
            continue
        # Total Operating Expenditures is col 3 (index 3) in Table 51
        try:
            total = float(r[3]) if r[3] is not None else None
        except (TypeError, ValueError):
            continue
        if not total or total <= 0:
            continue
        # Pad to 5 digits for state_leaid match
        try:
            padded = f"{int(code):05d}"
        except ValueError:
            continue
        out.append({
            "code": padded,
            "name": r[1],
            "total_op_exp": total,
        })
    return out


def build_tn_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("TN-"):
            out[sl.removeprefix("TN-")] = r
    return out


def extract(*, fiscal_year: int = 2025, triggered_by: str = "manual") -> dict:
    print(f"TN extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No TDOE ASR URL for fiscal_year={fiscal_year}; add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/',1)[-1]}...")
        zip_bytes = download(url)
        content_hash = sha256_bytes(zip_bytes)
        print(f"  {len(zip_bytes) / 1e6:.1f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/asr_excel.zip"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=zip_bytes,
                mime_type="application/zip",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/zip",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "ZIP contains 50+ Excel tables; use 'TABLE 51 24-25.xlsx'; "
                "match col 0 (3-digit, zero-pad to 5) == state_leaid suffix; "
                "topline = col 3 'TOTAL OPERATING EXPENDITURES'"
            ),
            notes=f"FY{fiscal_year} TDOE ASR (audited current expenditures)",
        )

        crosswalk = build_tn_crosswalk(client)
        print(f"  TN crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_data = parse_table51(zip_bytes)
        print(f"  ASR Table 51 districts: {len(district_data):,}")

        no_match: list[str] = []
        for d in district_data:
            district = crosswalk.get(d["code"])
            if district is None:
                no_match.append(d["code"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            _, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched ASR codes: {len(no_match)}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2025)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
