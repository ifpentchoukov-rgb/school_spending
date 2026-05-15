"""Arizona actuals extractor — ADE SAFR Digital Data (Districts + Charters).

Per A.R.S. § 15-901, AZ school districts and charter holders file an
Annual Financial Report (AFR) with the Arizona Department of Education's
School Finance Division each fall covering the prior FY (Jul 1–Jun 30).
ADE publishes the consolidated 'SAFR Digital Data' workbook each January
covering the full prior FY.

Source URL:
  https://www.azed.gov/sites/default/files/{YYYY}/01/
    Digital%20Data%20-%20Districts%20%26%20Charters%20Final.xlsx

Network note:
  azed.gov sits behind Akamai/Imperva and rejects Python's stdlib HTTP
  clients with browser-like headers. We use `curl_cffi` with
  `impersonate='chrome120'` to mimic a real Chrome TLS handshake — that
  passes the WAF cleanly. `verify=False` because curl-impersonate doesn't
  pick up macOS's cert bundle automatically; URL pin in source_documents
  preserves provenance.

Topline definition:
  Sum of all SAFR object columns across all 11 NCES function blocks
  (Function 1000 Instruction; 2100/2200/2300/2400/2500+2900/2600/2700
  Support Services; 3100 Food Service; 3200 Enterprise; 3400 Bookstore).
  Object codes: 6100 (salaries) + 6200 (benefits) + 6300/6400/6500
  (services) + 6600 (supplies) + 6810 (dues) + 6820 (charges) + 6841/2/3
  + 6850 (interest) + 6885/6890 (misc). This is exactly the F-33
  'current expenditures' frame: Function 4000 (Facilities) and Function
  5000 (Debt Service) are not included by SAFR's grid construction.

Status: `actual` — AFR is audited annual data.

Crosswalk:
  Master state_leaid format: 'AZ-{4-digit Entity ID}'
                              (e.g. 'AZ-4235' Mesa Unified District)
  SAFR file:                  Name + CTDS (no Entity ID directly)
  Match strategy:             normalize Name, strip the master's
                              ' (NNNN)' suffix, fuzzy-compare names.
"""

from __future__ import annotations

import argparse
import io
import re
import sys

import openpyxl
from curl_cffi import requests as curl_req
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "az"
STATE = "AZ"
BUCKET = "az"
SOURCE_PORTAL_URL = "https://www.azed.gov/finance"
PUBLISHER = "Arizona Department of Education (School Finance Division)"
DOCUMENT_TYPE = "ade_safr_digital_data_xlsx"
TOPLINE_DEFINITION = (
    "ADE SAFR Digital Data XLSX, sum of all object columns (6100 "
    "salaries + 6200 benefits + 6300/6400/6500 services + 6600 supplies "
    "+ 6810 dues + 6820 charges + 6841-50 interest + 6885/90 misc) "
    "across all 11 NCES function blocks (Function 1000 Instruction, "
    "2100-2900 Support Services, 3100/3200/3400 Operations). Excludes "
    "Function 4000 (Facilities Acquisition / Capital) and Function 5000 "
    "(Debt Service) by grid construction. F-33 'current expenditures' "
    "frame."
)
USER_AGENT = (
    "school-budget-tracker/0.1 "
    "(https://github.com/ifpentchoukov-rgb/school_spending)"
)

# Annual ADE publication URL — pinned per release year.
# Format: 'Digital Data - Districts & Charters Final.xlsx' under
# /sites/default/files/{YYYY}/01/ where YYYY is the publication
# calendar year (e.g. 2026/01 for FY25 data published Jan 2026).
KNOWN_FILE_URLS: dict[int, str] = {
    # FY25 (SY 2024-25) published Jan 2026
    2025: "https://www.azed.gov/sites/default/files/2026/01/Digital%20Data%20-%20Districts%20%26%20Charters%20Final.xlsx",
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    # Let curl-cffi use chrome120's native User-Agent — passing a custom
    # UA can break some WAFs (NH was 403'ing with one, OK without).
    r = curl_req.get(
        url,
        impersonate="chrome120",
        timeout=180,
        verify=False,
    )
    r.raise_for_status()
    return r.content


def _normalize_name(s: str) -> str:
    """Lowercase, strip ' (NNNN)' suffix, collapse whitespace, drop common
    suffix variations to enable fuzzy matching."""
    if not s:
        return ""
    s = s.strip()
    # Strip trailing ' (NNNN)' suffix from master
    s = re.sub(r"\s*\(\d{3,5}\)\s*$", "", s)
    s = s.lower()
    # Normalize "School District" / "District" / "Schools" — not used
    # consistently between SAFR and master:
    s = s.replace(" school district", " district")
    s = s.replace(" schools", "")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# Phase 7.5 — AZ SAFR function-block 0-indexed column ranges. Row 2 of
# both District and Charter sheets labels these spans.
# (1-indexed cols converted to 0-indexed half-open ranges below.)
_AZ_FUNCTION_BLOCKS: list[tuple[str, range]] = [
    ("instruction", range(3, 9)),            # F1000 Instruction (6 cols)
    ("support_services_student", range(9, 15)),       # F2100 (6)
    ("support_services_instruction", range(15, 21)),  # F2200 (6)
    ("administration", range(21, 28)),       # F2300 Gen Admin (7)
    ("administration", range(28, 34)),       # F2400 School Admin (6)
    ("administration", range(34, 41)),       # F2500/2900 Central + Other (7)
    ("operations_maintenance", range(41, 47)),        # F2600 (6)
    ("transportation", range(47, 53)),       # F2700 (6)
    ("food_service", range(53, 59)),         # F3100 (6)
    # F3200 Enterprise (cols 59-64) and F3400 Bookstore (65-70) intentionally
    # omitted from component breakdown — too small + no canonical match.
]
# Object 6200 (Employee Benefits) is at start+1 of each 6/7-col block.
_AZ_BENEFITS_COLS_0IDX: tuple[int, ...] = (
    4, 10, 16, 22, 29, 35, 42, 48, 54, 60, 66,
)


def parse_az_safr(xlsx_bytes: bytes) -> list[dict]:
    """Return [{name, ctds, total_op_exp, kind, components}] from both
    Districts and Charters sheets. Total = sum of all numeric cells from
    col D onwards (Function block columns); components decompose by
    Function block per _AZ_FUNCTION_BLOCKS plus Object 6200 benefits.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    out: list[dict] = []
    for sheet_name, kind in [("District", "district"), ("Charter", "charter")]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Header rows 1-6; data starts at row 7. Cols A=Fiscal Year,
        # B=Name, C=CTDS, D-onwards = function/object grid.
        for row in range(7, ws.max_row + 1):
            name = ws.cell(row, 2).value
            ctds = ws.cell(row, 3).value
            if not name:
                continue
            # Materialize row as list (1-indexed col→0-indexed-list value)
            row_vals: list[float | None] = []
            for col in range(4, ws.max_column + 1):
                v = ws.cell(row, col).value
                row_vals.append(v if isinstance(v, (int, float)) else None)
            total = sum(v for v in row_vals if v is not None)
            if total <= 0:
                continue
            # Phase 7.5 — canonical category breakdown. Note: row_vals is
            # 0-indexed starting at original col 4, so subtract 3 from each
            # 1-indexed col index used above.
            offset = 3  # column D (1-indexed 4) → row_vals[0]
            components: dict[str, float] = {}
            for category, col_range in _AZ_FUNCTION_BLOCKS:
                s = 0.0
                for c in col_range:
                    idx = c - offset
                    if 0 <= idx < len(row_vals) and row_vals[idx]:
                        s += row_vals[idx]
                if s > 0:
                    components[category] = components.get(category, 0.0) + s
            # employee_benefits = sum of Object 6200 cols across function blocks
            benefits = 0.0
            for c in _AZ_BENEFITS_COLS_0IDX:
                idx = c - offset
                if 0 <= idx < len(row_vals) and row_vals[idx]:
                    benefits += row_vals[idx]
            if benefits > 0:
                components["employee_benefits"] = benefits

            out.append({
                "name": str(name).strip(),
                "ctds": str(ctds).strip() if ctds else "",
                "total_op_exp": total,
                "kind": kind,
                "components": components,
            })
    return out


def build_az_crosswalk(client: Client) -> tuple[dict[str, dict], dict[str, dict]]:
    """Return (by_state_leaid, by_normalized_name) crosswalks."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    by_id: dict[str, dict] = {}
    by_norm: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("AZ-"):
            by_id[sl.removeprefix("AZ-")] = r
        norm = _normalize_name(r.get("lea_name") or "")
        if norm:
            by_norm[norm] = r
    return by_id, by_norm


def extract(*, fiscal_year: int = 2025, triggered_by: str = "manual") -> dict:
    print(f"AZ actuals extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No AZ SAFR URL for fiscal_year={fiscal_year}; "
                f"add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/', 1)[-1]} (curl-cffi chrome120)...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.2f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/safr_digital_data.xlsx"
        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheets 'District' + 'Charter'; sum cols D onwards "
                "per row (rows 7+); match LEA by normalized name "
                "(strip master suffix '(NNNN)')"
            ),
            notes=(
                f"FY{fiscal_year} AZ SAFR Digital Data (Districts + "
                f"Charters). Fetched via curl-cffi chrome120 to bypass "
                f"Akamai. F-33 frame: all functions 1000-3400; excludes "
                f"4000 capital + 5000 debt by grid construction."
            ),
        )

        by_id, by_norm = build_az_crosswalk(client)
        print(
            f"  AZ crosswalk: {len(by_id):,} by ID, "
            f"{len(by_norm):,} by normalized name"
        )

        records = parse_az_safr(xlsx_bytes)
        print(f"  SAFR records (districts + charters): {len(records):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for d in records:
            norm = _normalize_name(d["name"])
            district = by_norm.get(norm)
            if district is None:
                no_match.append(d["name"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            components: list[ComponentInput] = []
            for category, amount in d.get("components", {}).items():
                if amount <= 0:
                    continue
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=(
                            f"ADE SAFR Digital Data XLSX, {d['kind']} sheet, "
                            f"row matching name='{d['name']}' (CTDS={d['ctds']}): "
                            f"sum of object cols within function block(s) "
                            f"mapping to '{category}' (or Object 6200 across "
                            f"all blocks for employee_benefits)"
                        ),
                        line_or_cell_reference=(
                            f"Sheets District/Charter; row name={d['name']}; "
                            f"per _AZ_FUNCTION_BLOCKS / _AZ_BENEFITS_COLS"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched LEAs: {len(no_match)}"
        )
        if no_match[:5]:
            print(f"  sample unmatched: {no_match[:8]}")
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2025)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
