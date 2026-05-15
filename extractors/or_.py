"""Oregon extractor — ODE Detailed District Expenditure Data XLSX.

Source: https://www.oregon.gov/ode/schools-and-districts/fiscaltransparency/pages/district%20detailed%20revenue.aspx
File: {YYYY-YY}%20Actual%20Expenditure%20Data.xlsx — published annually
      by ODE Fiscal Transparency Unit; one row per district × school ×
      fund × function × object × area-of-responsibility tuple.

What this gives us:
  - Detail-level per-district expenditures across all funds and chart
    of accounts for ~184 OR operating LEAs. Sheet "{FY} Actual
    Expenditure Data" has 14 cols including FundCd, FunctionCd, and
    ActualExpAmt.

Topline definition:
  Sum of ActualExpAmt per Institution_ID where Function code's first
  digit is in {1, 2, 3} — Instruction (1XXX), Support Services (2XXX),
  Enterprise / Community Services (3XXX). Excludes Facilities
  Acquisition (4XXX) and Other Uses / Debt Service (5XXX). This in
  effect excludes Fund 300 (Debt Service) and Fund 400 (Capital
  Projects) since their spending lands in those function codes.
  Aligned with F-33 'current expenditures' frame.

Status: `actual` — post-AFR audited (ODE publishes ~12 months after
FY-end per their note).

Crosswalk:
  Master state_leaid format: 'OR-{14-digit}' zero-padded
                              (e.g. 'OR-00000000002193' Falls City SD)
  ODE Institution_ID:        4-digit integer (e.g. 2193)
  → strip leading zeros from state_leaid suffix == Institution_ID.
"""

from __future__ import annotations

import argparse
import io
import sys
import urllib.error
import urllib.request

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "or"
STATE = "OR"
BUCKET = "or"
SOURCE_PORTAL_URL = "https://www.oregon.gov/ode/schools-and-districts/fiscaltransparency/pages/district%20detailed%20revenue.aspx"
PUBLISHER = "Oregon Department of Education (Fiscal Transparency Unit)"
DOCUMENT_TYPE = "ode_actual_expenditure_xlsx"
TOPLINE_DEFINITION = (
    "ODE Detailed District Expenditure Data, sheet '{FY} Actual "
    "Expenditure Data' — sum of ActualExpAmt per Institution_ID where "
    "FunctionCd's first digit is in {1, 2, 3} (Instruction, Support "
    "Services, Enterprise/Community Services). Excludes Facilities "
    "Acquisition (4XXX) and Other Uses / Debt Service (5XXX). Aligned "
    "with F-33 'current expenditures' frame."
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"


def file_url(fiscal_year: int) -> str:
    end_yy = fiscal_year - 2000
    start_yy = end_yy - 1
    return (
        f"https://www.oregon.gov/ode/schools-and-districts/FiscalTransparency/"
        f"Documents/{2000+start_yy}-{end_yy:02d}%20Actual%20Expenditure%20Data.xlsx"
    )


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=180) as resp:
        return resp.read()


# Phase 7.5 — OR canonical category mapping by ODE function code (4-digit).
# Each category gets a callable that returns True for matching function codes.
def _or_func_to_category(fc: str) -> str | None:
    """Map a 4-digit FunctionCd to a canonical category. None means
    'leave out of component breakdown' (e.g. Community Services 3300)."""
    if not fc or not fc.isdigit() or len(fc) != 4:
        return None
    n = int(fc)
    if 1000 <= n <= 1999:
        return "instruction"
    if 2100 <= n <= 2199:
        return "support_services_student"
    if 2200 <= n <= 2299:
        return "support_services_instruction"
    if 2300 <= n <= 2399:
        return "administration"  # general admin
    if 2400 <= n <= 2499:
        return "administration"  # school admin (combined)
    if 2500 <= n <= 2519:
        return "administration"  # business services
    if 2540 <= n <= 2549:
        return "operations_maintenance"
    if 2550 <= n <= 2559:
        return "transportation"
    if 2520 <= n <= 2599:
        return "administration"  # other support — fold into admin
    if 3100 <= n <= 3199:
        return "food_service"
    if 3000 <= n <= 3099:
        return "food_service"  # food/enterprise lead range
    if 3200 <= n <= 3999:
        return None  # Community Services (3300), Enterprise (3500) — skip
    if 4000 <= n <= 4999:
        return "capital_outlay"
    if 5000 <= n <= 5999:
        return "debt_service"
    return None


def parse_or(xlsx_bytes: bytes, fiscal_year: int) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    # Sheet name pattern: "{YYYY-YY} Actual Expenditure Data"
    yy_label = f"{fiscal_year - 1}-{(fiscal_year - 2000):02d}"
    target = f"{yy_label} Actual Expenditure Data"
    if target not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{target}' not in workbook; got {wb.sheetnames}"
        )
    ws = wb[target]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    expected = (
        "SchoolYear", "Institution_ID", "Institution_Name", "School_InstID",
        "School_Name", "FundCd", "FundDesc", "FunctionCd",
    )
    if header[:8] != expected:
        raise RuntimeError(f"Unexpected ODE header: {header[:8]}")

    totals: dict[int, float] = {}
    components: dict[int, dict[str, float]] = {}
    for r in rows:
        if not r or r[1] is None or r[7] is None or r[13] is None:
            continue
        try:
            inst_id = int(r[1])
        except (TypeError, ValueError):
            continue
        try:
            amt = float(r[13])
        except (TypeError, ValueError):
            continue
        # Function code first digit (for topline)
        fc_str = str(r[7])
        fc_first = fc_str[0] if fc_str else ""
        if fc_first in {"1", "2", "3"}:
            totals[inst_id] = totals.get(inst_id, 0.0) + amt
        # Phase 7.5 — canonical category breakdown (includes 4XXX + 5XXX)
        category = _or_func_to_category(fc_str)
        if category is not None and amt > 0:
            components.setdefault(inst_id, {}).setdefault(category, 0.0)
            components[inst_id][category] += amt
    return [
        {
            "code": str(k),
            "total_op_exp": v,
            "components": components.get(k, {}),
        }
        for k, v in totals.items()
        if v > 0
    ]


def build_or_crosswalk(client: Client) -> dict[str, dict]:
    """Map Institution_ID (int as string, no leading zeros) → master row."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if not sl.startswith("OR-"):
            continue
        suffix = sl.removeprefix("OR-").lstrip("0")
        if suffix:
            out[suffix] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"OR extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        print(f"  downloading {url.rsplit('/', 1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.1f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/actual_expenditure_data.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet '{FY} Actual Expenditure Data'; group by Institution_ID; "
                "sum ActualExpAmt where FunctionCd[0] in (1,2,3); "
                "match Institution_ID == lstrip('0', state_leaid suffix)"
            ),
            notes=f"FY{fiscal_year} ODE Detailed District Expenditure Data",
        )

        crosswalk = build_or_crosswalk(client)
        print(f"  OR crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_data = parse_or(xlsx_bytes, fiscal_year=fiscal_year)
        print(f"  ODE districts with operating expenditures: {len(district_data):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for d in district_data:
            district = crosswalk.get(d["code"])
            if district is None:
                no_match.append(d["code"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            # Phase 7.5 — emit canonical category components.
            components: list[ComponentInput] = []
            for category, amount in d.get("components", {}).items():
                if amount <= 0:
                    continue
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=(
                            f"ODE Detailed District Expenditure Data: sum "
                            f"ActualExpAmt where Institution_ID={d['code']} "
                            f"AND FunctionCd maps to '{category}' per ODE "
                            f"Chart of Accounts function-range bucketing"
                        ),
                        line_or_cell_reference=(
                            f"Sheet '{fiscal_year-1}-{(fiscal_year-2000):02d} "
                            f"Actual Expenditure Data'; Institution_ID={d['code']}; "
                            f"function-code range bucketing per _or_func_to_category"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched ODE Institution_IDs: {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
