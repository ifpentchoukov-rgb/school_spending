"""Michigan extractor — MDE Bulletin 1011 (Analysis of MI Public Schools
Revenue and Expenditures) bulk Excel.

Source: https://www.michigan.gov/mde/services/financial-management/state-aid/
        publications/bulletin-1011-analysis-of-michigan-public-schools-
        revenue-and-expenditures
File pattern: https://mdoe.state.mi.us/SAMSPublic/Reports/others/
              {NN}_Bulletin1011Export.xlsx
e.g. 25_Bulletin1011Export.xlsx covers SY 2024-25 = our fiscal_year=2025.

What this gives us:
  - Per-district revenue + expenditure detail across 5 fund categories
    (General Fund, Special Revenue, Capital Projects, etc.) for completed
    Michigan FYs. MDE publishes one Excel annually after the AFR (Form
    SE-4096) reconciliation cycle completes.
  - 821 LEAs covered as of FY25.

Topline definition:
  Sum of `TOTCUROPEX` (Total Current Operating Expenditure) across all 5
  funds per district. This is MDE's all-funds operating spend, aligned
  with F-33 'current expenditures' and the actuals topline used for
  TX/CA/FL/IL/GA/OH.

Status: `actual` — these are post-AFR audited numbers.

Note: CEPI's Financial Information Database (FID) requires milogin and
isn't programmatically accessible. Bulletin 1011 is the public-facing
bulk extract of the same underlying data.

Crosswalk:
  Master state_leaid format: 'MI-{5-digit-DCode}' (e.g. 'MI-82015' Detroit)
  Bulletin DCode:            5-digit zero-padded code
  → strip 'MI-'.
"""

from __future__ import annotations

import argparse
import sys
import urllib.error
import urllib.request
from collections import defaultdict
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "mi"
STATE = "MI"
BUCKET = "mi"
SOURCE_PORTAL_URL = (
    "https://www.michigan.gov/mde/services/financial-management/state-aid/"
    "publications/bulletin-1011-analysis-of-michigan-public-schools-"
    "revenue-and-expenditures"
)
PUBLISHER = "Michigan Department of Education"
DOCUMENT_TYPE = "mde_bulletin_1011_xlsx"
TOPLINE_DEFINITION = (
    "MDE Bulletin 1011 Bulletin1011Export sheet, sum of TOTCUROPEX (Total "
    "Current Operating Expenditure) across all 5 funds per district — "
    "all-funds audited operating spend"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

# TOTCUROPEX column index in the Bulletin1011Export sheet (row 2 header).
TOTCUROPEX_COL_IDX = 59
DCODE_COL_IDX = 0
NAME_COL_IDX = 1

# Phase 7.4 canonical-category mapping. Each row in Bulletin 1011 is
# per-(district, fund); we sum each source column across all funds to
# get the per-district total for that category — consistent with the
# topline's all-funds sum. 11 of 14 canonical categories extractable.
# Omitted: support_services_student deliberately combined into the
# 'instructional staff' bucket (MDE doesn't separate counseling/health
# at the bulletin-export level); `other` left as residual.
#
# Source column legend (per MDE Form 6 / Bulletin 1011 documentation):
#   ITOT        Total Instruction (function 1XX)
#   TOTPUPSVC   Pupil services (functions 211–219) — counseling/health/etc.
#   TOTINSSTF   Instructional staff support (functions 221–229)
#   TOTSCHADM   School administration (functions 241/249)
#   TOTGENADM   General administration (functions 231/233/239)
#   TOTBUSADM   Business administration (functions 251–259)
#   TOTOPNMNT   Operations & maintenance (functions 261–269)
#   TOTTRANS   Transportation (functions 271–279)
#   SCHLUNCH    School lunch / food service
#   EMPBENINS   Employee benefits in instruction
#   TOTEBSUP    Total employee benefits in support
#   CAPOUTLAY   Capital outlay
#   DEBT        Debt service
#   LOCREV, STATEREV, FEDREV  Revenue by source
MI_COMPONENT_COLS: dict[str, tuple[list[int], str]] = {
    # category -> (col indexes to sum, definition fragment)
    "instruction": ([31], "MDE Bulletin 1011 ITOT (Total Instruction, function 1XX) summed across all funds"),
    "support_services_student": ([34], "MDE Bulletin 1011 TOTPUPSVC (Pupil Services — counseling, health, social work) summed across all funds"),
    "support_services_instruction": ([36], "MDE Bulletin 1011 TOTINSSTF (Instructional Staff Support — curriculum, media, in-service) summed across all funds"),
    "administration": (
        [38, 41, 43],
        "MDE Bulletin 1011 TOTSCHADM + TOTGENADM + TOTBUSADM summed across all funds",
    ),
    "operations_maintenance": ([47], "MDE Bulletin 1011 TOTOPNMNT (Operations & Maintenance) summed across all funds"),
    "transportation": ([49], "MDE Bulletin 1011 TOTTRANS (Pupil Transportation) summed across all funds"),
    "food_service": ([53], "MDE Bulletin 1011 SCHLUNCH (Food Service) summed across all funds"),
    "employee_benefits": (
        [29, 57],
        "MDE Bulletin 1011 EMPBENINS (benefits in instruction) + TOTEBSUP (benefits in support) summed across all funds — excludes benefits embedded in admin/transportation/food-service lines",
    ),
    "capital_outlay": ([62], "MDE Bulletin 1011 CAPOUTLAY summed across all funds"),
    "debt_service": ([66], "MDE Bulletin 1011 DEBT summed across all funds"),
    "revenue_federal": ([14], "MDE Bulletin 1011 FEDREV summed across all funds"),
    "revenue_state": ([13], "MDE Bulletin 1011 STATEREV summed across all funds"),
    "revenue_local": ([11], "MDE Bulletin 1011 LOCREV summed across all funds"),
}


def file_url(fiscal_year: int) -> str:
    nn = fiscal_year % 100
    return (
        f"https://mdoe.state.mi.us/SAMSPublic/Reports/others/"
        f"{nn:02d}_Bulletin1011Export.xlsx"
    )


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=120) as resp:
        return resp.read()


def parse_bulletin_1011(xlsx_bytes: bytes) -> list[dict]:
    """Aggregate TOTCUROPEX + canonical category breakdowns per district
    across all funds. Returns [{dcode, name, total_op_exp, components: {category: amount}}, ...]."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # Row 0 is blank, row 1 is header, data starts at row 2.
    header = rows[1] if len(rows) > 1 else None
    if not header or header[TOTCUROPEX_COL_IDX] != "TOTCUROPEX":
        raise RuntimeError(
            f"Header mismatch: expected TOTCUROPEX at col {TOTCUROPEX_COL_IDX}, "
            f"got '{header[TOTCUROPEX_COL_IDX] if header else None}'"
        )
    # Per-district totals: topline (TOTCUROPEX) and per-canonical-category.
    totals: dict[str, float] = defaultdict(float)
    names: dict[str, str] = {}
    component_totals: dict[str, dict[str, float]] = defaultdict(
        lambda: defaultdict(float)
    )
    for r in rows[2:]:
        if not r or not r[DCODE_COL_IDX]:
            continue
        try:
            v = float(r[TOTCUROPEX_COL_IDX] or 0)
        except (TypeError, ValueError):
            continue
        dcode = str(r[DCODE_COL_IDX])
        totals[dcode] += v
        names.setdefault(dcode, r[NAME_COL_IDX])

        # Sum canonical category source columns for this (district, fund)
        # row into the district's running per-category total.
        for category, (col_idxs, _def) in MI_COMPONENT_COLS.items():
            cat_amt = 0.0
            for ci in col_idxs:
                cv = r[ci] if ci < len(r) else None
                if cv is None:
                    continue
                try:
                    cat_amt += float(cv)
                except (TypeError, ValueError):
                    pass
            if cat_amt:
                component_totals[dcode][category] += cat_amt

    return [
        {
            "dcode": dcode,
            "name": names.get(dcode),
            "total_op_exp": total,
            "components": dict(component_totals[dcode]),
        }
        for dcode, total in totals.items()
        if total > 0
    ]


def build_mi_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("MI-"):
            out[sl.removeprefix("MI-")] = r
    return out


def extract(*, fiscal_year: int = 2025, triggered_by: str = "manual") -> dict:
    print(f"MI extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        print(f"  downloading {url.rsplit('/',1)[-1]}...")
        try:
            xlsx_bytes = download(url)
        except urllib.error.HTTPError as e:
            print(f"  FAILED: {e}")
            raise
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.1f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/bulletin_1011.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'Bulletin1011Export'; match DCode == state_leaid suffix; "
                "topline = sum(TOTCUROPEX) across all 5 funds per district"
            ),
            notes=f"FY{fiscal_year} Bulletin 1011 audited revenue/expenditure data",
        )

        crosswalk = build_mi_crosswalk(client)
        print(f"  MI crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_totals = parse_bulletin_1011(xlsx_bytes)
        print(f"  Bulletin 1011 districts: {len(district_totals):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for d in district_totals:
            district = crosswalk.get(d["dcode"])
            if district is None:
                no_match.append(d["dcode"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1

            # Phase 7.4 — emit canonical category components.
            components: list[ComponentInput] = []
            for category, amount in d.get("components", {}).items():
                if amount is None:
                    continue
                col_idxs, definition = MI_COMPONENT_COLS[category]
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=definition,
                        line_or_cell_reference=(
                            f"Sheet 'Bulletin1011Export'; sum cols {col_idxs} "
                            f"across all funds where DCode=={d['dcode']}"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch
            if changed:
                run.records_changed += 1

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched DCodes: {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "components_inserted": n_components_inserted,
        "components_updated": n_components_updated,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2025,
                   help="Bulletin 1011 FY (latest as of 2026-05-05: 2025)")
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
