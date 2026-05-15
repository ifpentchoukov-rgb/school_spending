"""Kentucky extractor — KDE AFR Revenues and Expenditures workbook.

Source: https://www.education.ky.gov/districts/FinRept/Pages/Fund%20Balances,%20Revenues%20and%20Expenditures,%20Chart%20of%20Accounts,%20Indirect%20Cost%20Rates%20and%20Key%20Financial%20Indicators.aspx
File: Revenues and Expenditures {YYYY-YY}.xlsx — KDE publishes annually,
      sourced from district MUNIS / Enterprise ERP submissions and audited
      Annual Financial Reports.

What this gives us:
  - Per-district expenditure detail by Function code (1000-5200) for all
    167 KY operating LEAs (county districts + independent districts).
  - Sheet '2024 AFR Expenditures ' (note trailing space) has columns by
    Function: Instruction (1000), Student Support (2100), Plant
    Operations (2600), etc., plus Facilities (4XXX), Debt Service (5100),
    Fund Transfers (5200), and pre-summed totals.

Topline definition:
  Sum of Function codes 1000-3900 per district — Instruction + Student
  Support + Instruction Staff + District Admin + School Admin + Business +
  Plant Operations + Pupil Transportation + Other Support + Food Service +
  Day Care + Community Services + Adult Education + Other Non-Instruction.
  Excludes Facilities (4XXX) and Debt Service (5100). Aligned with F-33
  'current expenditures' frame.

Status: `actual` — post-AFR audited.

Crosswalk:
  Master state_leaid format: 'KY-{9-digit}' where chars 3-5 are the
                              KDE district code (e.g. 'KY-001001000'
                              Adair County → KDE code '001').
  PDF/XLSX district label:    '{3-digit code} {District Name}' (e.g.
                              '001 Adair County').
  → state_leaid[3:6] == PDF 3-digit code.
"""

from __future__ import annotations

import argparse
import io
import re
import sys
import urllib.error
import urllib.request

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "ky"
STATE = "KY"
BUCKET = "ky"
SOURCE_PORTAL_URL = (
    "https://www.education.ky.gov/districts/FinRept/Pages/"
    "Fund%20Balances,%20Revenues%20and%20Expenditures,%20Chart%20of"
    "%20Accounts,%20Indirect%20Cost%20Rates%20and%20Key%20Financial"
    "%20Indicators.aspx"
)
PUBLISHER = "Kentucky Department of Education (Office of Finance and Operations)"
DOCUMENT_TYPE = "kde_afr_revenues_expenditures_xlsx"
TOPLINE_DEFINITION = (
    "KDE AFR Revenues and Expenditures workbook, '{YYYY} AFR "
    "Expenditures' sheet — sum of Function codes 1000-3900 per "
    "district (Instruction + Student Support + Plant Operations + "
    "Transportation + Food Service + Community Services + ...). "
    "Excludes Facilities 4XXX and Debt Service 5100. Aligned with "
    "F-33 'current expenditures' frame."
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

# Operating expense columns by index in the AFR Expenditures sheet
# (header in row 3): cols 2 through 15 inclusive cover Function 1000-3900.
OP_COL_START = 2  # Instruction 1000
OP_COL_END = 15  # Other Non-Instruction 3900

# Phase 7.4 — canonical category mapping for KY. Column indexes are
# 0-based against the AFR Expenditures sheet's data rows (row 3+).
# Each entry: canonical category -> (list of col indexes, definition fragment).
KY_COMPONENT_COLS: dict[str, tuple[list[int], str]] = {
    "instruction": ([2], "KDE AFR Function 1000 (Instruction)"),
    "support_services_student": ([3], "KDE AFR Function 2100 (Student Support)"),
    "support_services_instruction": ([4], "KDE AFR Function 2200 (Instruction Staff)"),
    "administration": (
        [5, 6, 7],
        "KDE AFR Functions 2300 (District Admin) + 2400 (School Admin) + 2500 (Business)",
    ),
    "operations_maintenance": ([8], "KDE AFR Function 2600 (Plant Operations)"),
    "transportation": ([9], "KDE AFR Function 2700 (Pupil Transportation)"),
    "food_service": ([11], "KDE AFR Function 3100 (Food Service)"),
    "capital_outlay": (
        [16, 17, 18, 19, 20, 21, 22, 23],
        "KDE AFR Function 4100-4900 (Facilities — Land Acquisition, Improvements, "
        "Architecture, Building Construction/Improvement, Other Facilities Acquisition)",
    ),
    "debt_service": ([24], "KDE AFR Function 5100 (Debt Service)"),
}


def file_url(fiscal_year: int) -> str:
    end_yy = fiscal_year - 2000  # 2024 -> 24
    start_yy = end_yy - 1  # 2023
    return (
        f"https://www.education.ky.gov/districts/FinRept/Documents/"
        f"Revenues%20and%20Expenditures%20{2000+start_yy}-{2000+end_yy}.xlsx"
    )


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


_CODE_RE = re.compile(r"^\s*(\d{3})\s+(.+?)\s*$")


def parse_ky(xlsx_bytes: bytes, fiscal_year: int) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    # Sheet name has a trailing space in the published file:
    target = f"{fiscal_year} AFR Expenditures "
    if target not in wb.sheetnames:
        # Fallback to no-trailing-space form
        alt = f"{fiscal_year} AFR Expenditures"
        if alt in wb.sheetnames:
            target = alt
        else:
            raise RuntimeError(
                f"Expected sheet '{target}' (or no trailing space) in KDE "
                f"workbook; got {wb.sheetnames}"
            )
    ws = wb[target]
    out: list[dict] = []
    rows = ws.iter_rows(values_only=True)
    # Row 0 = "Expenditures", row 1 = "For Fiscal Year YYYY",
    # row 2 = column headers, row 3+ = data.
    next(rows)
    next(rows)
    next(rows)
    for r in rows:
        if not r or r[0] is None:
            continue
        m = _CODE_RE.match(str(r[0]))
        if not m:
            continue
        code = m.group(1)
        # Sum operating cols (handle None / blank)
        total = 0.0
        for c in range(OP_COL_START, OP_COL_END + 1):
            if c >= len(r):
                break
            v = r[c]
            if v is None or v == "":
                continue
            try:
                total += float(v)
            except (TypeError, ValueError):
                continue
        if total <= 0:
            continue

        # Phase 7.4 — canonical category breakdown
        components: dict[str, float] = {}
        for category, (col_idxs, _def) in KY_COMPONENT_COLS.items():
            cat_total = 0.0
            for ci in col_idxs:
                if ci >= len(r):
                    continue
                v = r[ci]
                if v is None or v == "":
                    continue
                try:
                    cat_total += float(v)
                except (TypeError, ValueError):
                    continue
            if cat_total > 0:
                components[category] = cat_total

        out.append({
            "code": code,
            "name": m.group(2),
            "total_op_exp": total,
            "components": components,
        })
    return out


def build_ky_crosswalk(client: Client) -> dict[str, dict]:
    """Master state_leaid is 'KY-{9-digit}'; KDE district code = chars 3:6
    (0-indexed) of the 9-digit suffix."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if not sl.startswith("KY-"):
            continue
        suffix = sl.removeprefix("KY-")
        if len(suffix) >= 6:
            kde_code = suffix[3:6]
            out[kde_code] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"KY extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        print(f"  downloading {url.rsplit('/', 1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.2f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/afr_revenues_expenditures.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet '{FY} AFR Expenditures '; sum cols 2-15 (Function "
                "1000-3900) per row; match leading 3-digit code in "
                "District column == chars 3-5 of state_leaid 9-digit suffix"
            ),
            notes=f"FY{fiscal_year} KDE AFR Revenues and Expenditures",
        )

        crosswalk = build_ky_crosswalk(client)
        print(f"  KY crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_data = parse_ky(xlsx_bytes, fiscal_year=fiscal_year)
        print(f"  KDE districts with FY{fiscal_year} expenditures: {len(district_data):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for d in district_data:
            district = crosswalk.get(d["code"])
            if district is None:
                no_match.append(f"{d['code']} {d['name']}")
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            # Phase 7.4 — emit canonical category components.
            components: list[ComponentInput] = []
            for category, amount in d.get("components", {}).items():
                col_idxs, definition = KY_COMPONENT_COLS[category]
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=definition,
                        line_or_cell_reference=(
                            f"Sheet '{fiscal_year} AFR Expenditures '; "
                            f"sum cols {col_idxs} on row for district '{d['code']} {d['name']}'"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched KDE codes: {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
