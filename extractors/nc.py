"""North Carolina extractor — NCDPI State Public School Fund (SPSF) expenditures.

Source: https://www.dpi.nc.gov/districts-schools/district-operations/
        financial-and-business-services/demographics-and-finances
File pattern (FY25):
  https://www.dpi.nc.gov/documents/fbs/resources/
    fy{YYYY}spsfbyleabyprcplainenglish-rptxlsx/download?attachment
e.g. fy2025spsfbyleabyprcplainenglish-rptxlsx for SY 2024-25.

What this gives us:
  - **State-funded** operating expenditures per LEA per Program Report Code
    (PRC). NCDPI publishes one Excel annually around fall after FY close.
  - 115 county/city school district LEAs covered (NC has 116 traditional
    LEAs; charter LEAs file separately and are NOT in this file).

⚠️ TOPLINE LIMITATION:
  This is **state-funded only** — approximately 55-60% of a typical NC
  district's total operating budget. State funds + local appropriation +
  federal funds = total operating spend. The full all-funds figure
  requires the LGC audit data which is per-district PDF only (queued as
  a future Phase 6 follow-up extractor).

  When comparing to TX/CA/FL/IL/GA/OH actuals (which are total operating
  spend), NC will appear smaller than its actual size. Verifiers and
  downstream rollups should know this.

Topline definition (state-funded only):
  Sum of YTDExpenditures across all PRCs for each LEA from the
  'Data Tables' sheet's Key (PRC-LEA) column.

Status: `actual` — these are post-audit numbers (NC FY ends June 30,
report published fall of following year).

Crosswalk:
  Master state_leaid format: 'NC-{3-digit-LEA}' (e.g. 'NC-920' Wake)
  SPSF Key format:           'PRC-LEA' (e.g. '001-920' for PRC 1 Wake)
  → strip 'NC-' from master; split SPSF Key on '-' and take the LEA half.
"""

from __future__ import annotations

import argparse
import sys
import urllib.error
import urllib.request
from collections import defaultdict
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "nc"
STATE = "NC"
BUCKET = "nc"
SOURCE_PORTAL_URL = (
    "https://www.dpi.nc.gov/districts-schools/district-operations/"
    "financial-and-business-services/demographics-and-finances"
)
PUBLISHER = "North Carolina Department of Public Instruction"
DOCUMENT_TYPE = "ncdpi_spsf_xlsx"
TOPLINE_DEFINITION = (
    "NCDPI SPSF (State Public School Fund) operating expenditures per LEA "
    "— STATE-FUNDED ONLY (~55-60% of total operating budget). Sum of "
    "YTDExpenditures across all PRCs from the 'Data Tables' sheet."
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

# Path varies slightly by year (sometimes /resources/, sometimes /finance/).
# Hard-code known URLs; fall back to constructing the path.
KNOWN_FILE_URLS: dict[int, str] = {
    2025: "https://www.dpi.nc.gov/documents/fbs/resources/fy2025spsfbyleabyprcplainenglish-rptxlsx/download?attachment",
    2024: "https://www.dpi.nc.gov/documents/fbs/finance/fy2024spsfbyleabyprcplainenglish-rptxlsx/download?attachment",
    2023: "https://www.dpi.nc.gov/documents/fbs/finance/reporting/fy2023spsfbyleabyprcplainenglish-rptxlsx/download?attachment",
}


def discover_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=120) as resp:
        return resp.read()


def parse_data_tables(xlsx_bytes: bytes) -> dict[str, float]:
    """Aggregate state-funded YTD expenditures per LEA. Returns {lea_code: total}.
    Skips the 'ALL' synthetic row (statewide total)."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    if "Data Tables" not in wb.sheetnames:
        raise RuntimeError(f"Data Tables sheet missing; sheets={wb.sheetnames}")
    ws = wb["Data Tables"]
    totals: dict[str, float] = defaultdict(float)
    for r in range(4, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        if not key or not isinstance(key, str) or "-" not in key:
            continue
        parts = key.split("-")
        if len(parts) != 2:
            continue
        prc, lea = parts
        if lea == "ALL" or prc == "ALL":
            continue
        try:
            totals[lea] += float(val or 0)
        except (TypeError, ValueError):
            continue
    return dict(totals)


def build_nc_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("NC-"):
            out[sl.removeprefix("NC-")] = r
    return out


def extract(*, fiscal_year: int = 2025, triggered_by: str = "manual") -> dict:
    print(f"NC extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = discover_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No NC SPSF URL for fiscal_year={fiscal_year}; add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/',2)[-2]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.1f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/spsf.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'Data Tables', column A 'Key' (PRC-LEA), column B "
                "'YTDExpenditures'; aggregate by LEA suffix; "
                "match to state_leaid suffix"
            ),
            notes=(
                f"FY{fiscal_year} STATE-FUNDED expenditures only "
                "(~55-60% of total operating; charter LEAs not included)"
            ),
        )

        crosswalk = build_nc_crosswalk(client)
        print(f"  NC crosswalk: {len(crosswalk):,} state→NCES mappings")

        totals = parse_data_tables(xlsx_bytes)
        print(f"  SPSF LEAs aggregated: {len(totals):,}")

        no_match: list[str] = []
        for lea_code, total in sorted(totals.items()):
            district = crosswalk.get(lea_code)
            if district is None:
                no_match.append(lea_code)
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=total,
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            _, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched LEA codes: {len(no_match)}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2025)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
