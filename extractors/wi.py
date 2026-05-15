"""Wisconsin extractor — DPI Comparative Cost Per Member summary file.

Source: https://dpi.wi.gov/sfs/statistical/cost-revenue/section-d
File: compcost_sum_{YYYY}_to_{YYYY}_{date-suffix}.xlsx
      (Single multi-year XLSX; covers FY2008-09 through latest published.
      File name's date suffix changes when DPI republishes — pin per FY.)

What this gives us:
  - Per-district operating expenditures for all 377+ WI public school
    districts. DPI publishes this as the official "Comparative Cost
    Per Member" summary, derived from WUFAR Annual Report submissions.
  - DATA sheet has one row per district with cost columns repeated
    per fiscal year (instruct + support + admin + operations + trans +
    facility + food).

Topline definition:
  Sum of seven per-FY cost columns: instruct + support + admin +
  operations + trans + facility + food. Aligned with F-33 'current
  expenditures' frame; matches DPI's published "Total Cost" used in
  Comparative Cost Per Member calculations. Excludes debt service
  funds, capital projects, and inter-fund transfers.

Status: `actual` — post-WUFAR audited.

Crosswalk:
  Master state_leaid format: 'WI-{4-digit}' (zero-padded, e.g. 'WI-0007')
  DPI CODE column:           integer (e.g. 7)
  → zfill(CODE, 4) == state_leaid suffix.
"""

from __future__ import annotations

import argparse
import io
import sys
import urllib.error
import urllib.request

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "wi"
STATE = "WI"
BUCKET = "wi"
SOURCE_PORTAL_URL = "https://dpi.wi.gov/sfs/statistical/cost-revenue/section-d"
PUBLISHER = "Wisconsin Department of Public Instruction (School Financial Services)"
DOCUMENT_TYPE = "wi_dpi_compcost_summary_xlsx"
TOPLINE_DEFINITION = (
    "DPI Comparative Cost Per Member summary, DATA sheet — sum of "
    "seven per-FY cost columns: instruct + support + admin + "
    "operations + trans + facility + food. Aligned with F-33 'current "
    "expenditures' frame; excludes debt service, capital projects, "
    "and inter-fund transfers."
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

KNOWN_FILE_URLS: dict[int, str] = {
    # FY24 (SY 2023-24) — date suffix from file's last republish; bump when DPI updates.
    2024: (
        "https://dpi.wi.gov/sites/default/files/imce/sfs/xls/"
        "compcost_sum_0809_to_2324_20260316.xlsx"
    ),
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def _find_fy_block(header: tuple, target_fy: int, ws) -> int:
    """Find the column index where row 2's value == target_fy.

    The header row (row 0) repeats 'fiscal_year' (or older capitalization)
    every 8 cols; the actual numeric FY value is in data rows. Use
    Abbotsford row (row 2 by zero-index, the third row of the sheet) as
    the FY-locator.
    """
    locator = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    for i, val in enumerate(locator):
        if val == target_fy:
            return i
    raise RuntimeError(f"FY {target_fy} not found in DATA sheet locator row")


# Phase 7.4 — DPI Comparative Cost cost-column positions within an FY
# block (8 cols per FY: fiscal_year, member, instruct, support, admin,
# operations, trans, facility, food). Mapping cost-column offset →
# canonical category.
WI_COMPONENT_OFFSETS: dict[str, tuple[int, str]] = {
    "instruction": (
        2,
        "DPI Comparative Cost 'instruct' column — instruction (WUFAR Function 100000)",
    ),
    "support_services_student": (
        3,
        "DPI Comparative Cost 'support' column — combined pupil + instructional staff "
        "support (WUFAR Functions 210000+220000); WI does not separate the two",
    ),
    "administration": (
        4,
        "DPI Comparative Cost 'admin' column — combined district + school admin + "
        "business (WUFAR Functions 230000+240000+250000)",
    ),
    "operations_maintenance": (
        5,
        "DPI Comparative Cost 'operations' column — operation & maintenance of plant "
        "(WUFAR Function 260000)",
    ),
    "transportation": (
        6,
        "DPI Comparative Cost 'trans' column — pupil transportation "
        "(WUFAR Function 270000)",
    ),
    "capital_outlay": (
        7,
        "DPI Comparative Cost 'facility' column — facilities acquisition + "
        "improvements (WUFAR Function 280000)",
    ),
    "food_service": (
        8,
        "DPI Comparative Cost 'food' column — food service operations "
        "(WUFAR Function 410000 / 420000)",
    ),
}


def parse_compcost(xlsx_bytes: bytes, fiscal_year: int) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["DATA"]
    header = next(ws.iter_rows(values_only=True))
    fy_col = _find_fy_block(header, fiscal_year, ws)
    # 8 cols per FY block: [fiscal_year, member, instruct, support, admin,
    # operations, trans, facility, food]. Cost cols are fy_col+2 .. fy_col+8.
    cost_cols = list(range(fy_col + 2, fy_col + 9))

    out: list[dict] = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[0] in (None, "", "CODE"):
            continue
        try:
            code = int(r[0])
        except (TypeError, ValueError):
            continue
        # Confirm this row's FY matches (some rows may be sparse)
        if r[fy_col] != fiscal_year:
            continue
        total = 0.0
        for c in cost_cols:
            v = r[c]
            if v is None:
                continue
            try:
                total += float(v)
            except (TypeError, ValueError):
                continue
        if total <= 0:
            continue

        # Phase 7.4 — canonical category breakdown
        components: dict[str, float] = {}
        for category, (offset, _def) in WI_COMPONENT_OFFSETS.items():
            cv = r[fy_col + offset] if fy_col + offset < len(r) else None
            if cv is None:
                continue
            try:
                amt = float(cv)
            except (TypeError, ValueError):
                continue
            if amt > 0:
                components[category] = amt

        out.append({
            "code": f"{code:04d}",
            "total_op_exp": total,
            "components": components,
            "fy_col": fy_col,
        })
    return out


def build_wi_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("WI-"):
            out[sl.removeprefix("WI-")] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"WI extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No DPI compcost URL for fiscal_year={fiscal_year}; "
                "add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/', 1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.2f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/dpi_compcost_summary.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'DATA'; locate column where row 3 has FY=N; sum 7 "
                "cost columns (instruct, support, admin, operations, trans, "
                "facility, food); match zfill(CODE, 4) == state_leaid suffix"
            ),
            notes=f"FY{fiscal_year} DPI Comparative Cost Per Member summary",
        )

        crosswalk = build_wi_crosswalk(client)
        print(f"  WI crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_data = parse_compcost(xlsx_bytes, fiscal_year=fiscal_year)
        print(f"  DPI districts with FY{fiscal_year} cost data: {len(district_data):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for d in district_data:
            district = crosswalk.get(d["code"])
            if district is None:
                no_match.append(d["code"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            # Phase 7.4 — emit canonical category components.
            components: list[ComponentInput] = []
            for category, amount in d.get("components", {}).items():
                offset, definition = WI_COMPONENT_OFFSETS[category]
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=definition,
                        line_or_cell_reference=(
                            f"Sheet 'DATA'; cell at FY-block col {d['fy_col']}+{offset} "
                            f"on row with CODE={d['code'].lstrip('0') or '0'}"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched DPI codes: {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
