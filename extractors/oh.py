"""Ohio extractor — ODE District Profile Report ("Cupp Report") bulk Excel.

Source: https://education.ohio.gov/Topics/Finance-and-Funding/School-Payment-Reports/
        District-Profile-Reports
File pattern (post-2026 versioning):
  https://education.ohio.gov/getattachment/Topics/Finance-and-Funding/
    School-Payment-Reports/District-Profile-Reports/
    FY{NN}-District-Profile-Report/
    FY{NN}-District-Profile-Report-Final-Revised-{M-DD-YY}-posted.xlsx.aspx?lang=en-US

Filename includes a "posted" date that's not predictable, so we hard-code
the latest URL and add an index-page-scrape fallback.

What this gives us:
  - Per-district demographics, valuation, millage, expenditure-per-pupil,
    and revenue breakdown for completed Ohio FYs. ODE publishes one Excel
    annually around March (FY25 was posted 3-10-26).

Topline definition:
  Computed: `Enrolled ADM FY{NN}` × `Total Operating Expenditure Per Pupil
  FY{NN}`. The Cupp Report doesn't expose absolute total spend; it
  expresses operating expense as a per-pupil ratio. Multiplying by ADM
  reconstructs the total. Aligned with F-33's 'current expenditures'
  concept — comparable to the actuals topline used for TX/CA/FL/IL/GA.

Status: `actual` — these are post-audit numbers from the underlying
USAS/EMIS data ODE compiles.

Crosswalk:
  Master state_leaid format: 'OH-{6-digit-IRN}' (e.g. 'OH-043489' Akron)
  Cupp IRN column:           6-digit zero-padded IRN (e.g. '043489')
  → strip 'OH-'.
"""

from __future__ import annotations

import argparse
import re
import sys
import urllib.error
import urllib.request
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "oh"
STATE = "OH"
BUCKET = "oh"
SOURCE_PORTAL_URL = (
    "https://education.ohio.gov/Topics/Finance-and-Funding/"
    "School-Payment-Reports/District-Profile-Reports"
)
PUBLISHER = "Ohio Department of Education and Workforce"
DOCUMENT_TYPE = "ode_cupp_report_xlsx"
TOPLINE_DEFINITION = (
    "ODE District Profile Report (Cupp), 'District Data' sheet: "
    "Enrolled ADM FY{NN} × Total Operating Expenditure Per Pupil FY{NN} "
    "= total operating expenditure per district (audited actual)"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

# Filename includes a "posted" date that's not predictable. Cache known good
# URLs per FY; fall back to scraping the per-FY page for the latest matching
# attachment.
KNOWN_FILE_URLS: dict[int, str] = {
    2025: (
        "https://education.ohio.gov/getattachment/Topics/Finance-and-Funding/"
        "School-Payment-Reports/District-Profile-Reports/"
        "FY2025-District-Profile-Report/"
        "FY25-District-Profile-Report-Final-Revised-3-10-26-posted.xlsx.aspx"
        "?lang=en-US"
    ),
}

INDEX_URL_TEMPLATE = (
    "https://education.ohio.gov/Topics/Finance-and-Funding/School-Payment-Reports/"
    "District-Profile-Reports/FY{fy}-District-Profile-Report"
)


def discover_url(fiscal_year: int) -> str | None:
    if fiscal_year in KNOWN_FILE_URLS:
        return KNOWN_FILE_URLS[fiscal_year]
    index_url = INDEX_URL_TEMPLATE.format(fy=fiscal_year)
    req = urllib.request.Request(index_url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
    except urllib.error.URLError:
        return None
    pattern = re.compile(
        r'href="(/getattachment/[^"]*FY' + str(fiscal_year)[2:] +
        r'-District-Profile-Report-[^"]+\.xlsx\.aspx[^"]*)"'
    )
    m = pattern.search(html)
    if not m:
        return None
    return f"https://education.ohio.gov{m.group(1)}"


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=120) as resp:
        return resp.read()


# Phase 7.4 — Cupp Report per-pupil category columns. Each label
# template uses FY-short suffix; multiplied by ADM to reconstruct total
# dollars. 8 canonical categories: 5 expenditure + 3 revenue.
OH_PERPUPIL_LABELS: dict[str, str] = {
    # category -> Cupp column label template (with FY-short suffix appended)
    "instruction": "Instructional Expenditure Per Pupil FY{fy}",
    "support_services_student": "Pupil Support Expenditure Per Pupil FY{fy}",
    "support_services_instruction": "Staff Support Expenditure Per Pupil FY{fy}",
    "administration": "Administrator Expenditure Per Pupil FY{fy}",
    "operations_maintenance": "Building Operation Expenditure Per Pupil FY{fy}",
    "revenue_state": "State Revenue Per Pupil FY{fy}",
    "revenue_local": "Local Revenue Per Pupil FY{fy}",
    "revenue_federal": "Federal Revenue Per Pupil FY{fy}",
}


def parse_district_data(xlsx_bytes: bytes, fiscal_year: int) -> list[dict]:
    """Return [{irn, name, adm, oepp, total_op_exp, perpupil_components}, ...]."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "District Data" not in wb.sheetnames:
        raise RuntimeError(f"District Data sheet missing; sheets={wb.sheetnames}")
    ws = wb["District Data"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    fy_short = str(fiscal_year)[2:]  # 2025 → '25'
    adm_label = f"Enrolled ADM FY{fy_short}"
    oepp_label = f"Total Operating Expenditure Per Pupil FY{fy_short}"
    try:
        adm_col = header.index(adm_label)
        oepp_col = header.index(oepp_label)
    except ValueError:
        raise RuntimeError(
            f"Expected columns not found ({adm_label!r}, {oepp_label!r}). "
            f"Header sample: {header[:8]} ..."
        )
    irn_col = header.index("IRN") if "IRN" in header else 1
    name_col = header.index("District") if "District" in header else 0

    # Phase 7.4 — find per-pupil component columns
    pp_cols: dict[str, int] = {}
    for category, label_tpl in OH_PERPUPIL_LABELS.items():
        label = label_tpl.format(fy=fy_short)
        if label in header:
            pp_cols[category] = header.index(label)

    out: list[dict] = []
    for r in rows[1:]:
        if not r or not r[irn_col]:
            continue
        try:
            adm = float(r[adm_col]) if r[adm_col] is not None else None
            oepp = float(r[oepp_col]) if r[oepp_col] is not None else None
        except (TypeError, ValueError):
            continue
        if not adm or not oepp:
            continue
        # Phase 7.4 — per-pupil component values
        pp_components: dict[str, float] = {}
        for category, col_idx in pp_cols.items():
            v = r[col_idx] if col_idx < len(r) else None
            if v is None:
                continue
            try:
                amt = float(v)
            except (TypeError, ValueError):
                continue
            if amt > 0:
                pp_components[category] = amt
        out.append({
            "irn": str(r[irn_col]),
            "name": r[name_col],
            "adm": adm,
            "oepp": oepp,
            "total_op_exp": adm * oepp,
            "perpupil_components": pp_components,
        })
    return out


def build_oh_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("OH-"):
            out[sl.removeprefix("OH-")] = r
    return out


def extract(*, fiscal_year: int = 2025, triggered_by: str = "manual") -> dict:
    print(f"OH extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = discover_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No Cupp Report URL for fiscal_year={fiscal_year}; add to "
                "KNOWN_FILE_URLS or check that the FY page is published."
            )
        print(f"  downloading {url.rsplit('/',1)[-1].split('?')[0]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.1f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/cupp_report.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                f"Sheet 'District Data'; match IRN == state_leaid suffix; "
                f"topline = `Enrolled ADM FY{str(fiscal_year)[2:]}` × "
                f"`Total Operating Expenditure Per Pupil FY{str(fiscal_year)[2:]}`"
            ),
            notes=(
                f"FY{fiscal_year} Cupp Report (audited per-pupil + ADM)"
            ),
        )

        crosswalk = build_oh_crosswalk(client)
        print(f"  OH crosswalk: {len(crosswalk):,} state→NCES mappings")

        cupp_rows = parse_district_data(xlsx_bytes, fiscal_year)
        print(f"  Cupp District Data rows: {len(cupp_rows):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        fy_short = str(fiscal_year)[2:]
        for row in cupp_rows:
            district = crosswalk.get(row["irn"])
            if district is None:
                no_match.append(row["irn"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=row["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION.replace("{NN}", fy_short),
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            # Phase 7.4 — emit per-pupil × ADM components.
            adm = row["adm"]
            components: list[ComponentInput] = []
            for category, per_pupil in row.get("perpupil_components", {}).items():
                label = OH_PERPUPIL_LABELS[category].format(fy=fy_short)
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(per_pupil) * float(adm),
                        definition=(
                            f"ODE Cupp Report 'District Data' sheet, "
                            f"column '{label}' × 'Enrolled ADM FY{fy_short}'"
                        ),
                        line_or_cell_reference=(
                            f"Row for IRN={row['irn']}; column '{label}'"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched IRNs: {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2025,
                   help="Cupp Report FY (latest as of 2026-05-05: 2025)")
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
