"""Nebraska actuals extractor — NE SFOS Annual Financial Report (AFR).

Per NRS §79-1095, NE school districts and ESUs file an Annual Financial
Report (AFR) with the Nebraska Department of Education's School Finance
and Organization Services (SFOS) office. SFOS publishes consolidated
per-district AFR data each fall as a single XLSX inside a ZIP.

Source URL pattern (FY25 example):
  https://sfos.education.ne.gov/FOS/Data/afr{YYYY-2}{YYYY}.zip
  e.g. afr2425.zip = SY 2024-25 = our fiscal_year=2025

The original deferral note ('sfos.education.ne.gov ASP.NET per-district
interactive') was wrong — sfos.education.ne.gov/Default.aspx is a
static HTML page that links directly to per-FY ZIP archives. No
postback flow needed.

Topline definition:
  Account `01-2-20400-000` 'TOTAL GENERAL FUND EXPENDITURES' per
  district. NE's headline operating figure — General Fund only,
  excluding capital outlay (Funds 02/07/08), debt service (Fund 06),
  cooperative (04), activity (05), nutrition (10), and student fee (12)
  funds. Narrower than F-33 'current expenditures' but the most
  apples-to-apples figure NE publishes; the file also has Total
  Disbursements (01-2-20500-000, all-funds) for cross-checking.

Status: `actual` — AFR is the audited annual filing.

Crosswalk:
  Master state_leaid format: 'NE-{9-digit no-dash AgencyID}'
                              (e.g. 'NE-550001000' Lincoln Public Schools)
  XLSX AgencyID column:       formatted '55-0001-000' (with dashes)
  → state_leaid suffix == AgencyID.replace('-', '').
"""

from __future__ import annotations

import argparse
import io
import sys
import zipfile

import openpyxl
from curl_cffi import requests as curl_req
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "ne"
STATE = "NE"
BUCKET = "ne"
SOURCE_PORTAL_URL = "https://sfos.education.ne.gov/Default.aspx"
PUBLISHER = (
    "Nebraska Department of Education "
    "(School Finance and Organization Services)"
)
DOCUMENT_TYPE = "ne_sfos_afr_xlsx"
TOPLINE_DEFINITION = (
    "NE SFOS AFR XLSX — Account '01-2-20400-000' 'TOTAL GENERAL FUND "
    "EXPENDITURES' per AgencyID. General Fund operating only; "
    "excludes capital (Funds 02/07/08), debt service (Fund 06), "
    "and cooperative/activity/nutrition/student-fee funds (04, 05, "
    "09, 10, 12). Narrower than F-33 'current expenditures' but the "
    "headline operating figure NE publishes."
)

# General Fund Total Expenditures account (operating frame).
TOPLINE_ACCOUNT = "01-2-20400-000"


def _afr_url(fiscal_year: int) -> str:
    """e.g. fiscal_year=2025 -> afr2425.zip (SY 2024-25)."""
    yy1 = (fiscal_year - 1) % 100
    yy2 = fiscal_year % 100
    return (
        f"https://sfos.education.ne.gov/FOS/Data/"
        f"afr{yy1:02d}{yy2:02d}.zip"
    )


KNOWN_FILE_URLS: dict[int, str] = {
    # FY25 = SY 2024-25 (latest as of 2026-05-07).
    2025: _afr_url(2025),
    # FY24 = SY 2023-24 backfill.
    2024: _afr_url(2024),
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year, _afr_url(fiscal_year))


def download(url: str) -> bytes:
    r = curl_req.get(
        url,
        impersonate="chrome120",
        timeout=120,
        verify=False,
    )
    r.raise_for_status()
    return r.content


def parse_ne_afr(zip_bytes: bytes) -> list[dict]:
    """Return [{agency_id, total_op_exp}] from the AFR ZIP."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = [n for n in zf.namelist() if n.endswith(".xlsx")]
        if not names:
            raise RuntimeError(
                f"No .xlsx inside ZIP. Contents: {zf.namelist()}"
            )
        with zf.open(names[0]) as f:
            xlsx_bytes = f.read()

    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), data_only=True, read_only=True
    )
    if "AFR" not in wb.sheetnames:
        raise RuntimeError(
            f"Expected sheet 'AFR' not found. Sheets: {wb.sheetnames}"
        )
    ws = wb["AFR"]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    try:
        i_agency = header.index("AgencyID")
        i_acct = header.index("Account")
        i_amt = header.index("Amount")
    except ValueError as e:
        raise RuntimeError(f"Missing AFR column: {e}") from e

    out: list[dict] = []
    for row in rows:
        if not row or row[i_agency] is None or row[i_acct] is None:
            continue
        if str(row[i_acct]).strip() != TOPLINE_ACCOUNT:
            continue
        try:
            amt = float(row[i_amt] or 0)
        except (TypeError, ValueError):
            continue
        if amt <= 0:
            continue
        agency = str(row[i_agency]).strip()
        out.append({
            "agency_id": agency,
            "total_op_exp": amt,
        })
    return out


def build_ne_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("NE-"):
            out[sl.removeprefix("NE-").strip()] = r
    return out


def extract(*, fiscal_year: int = 2025, triggered_by: str = "manual") -> dict:
    print(f"NE actuals extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No NE AFR URL for fiscal_year={fiscal_year}."
            )
        print(f"  downloading {url.rsplit('/', 1)[-1]} (curl-cffi chrome120)...")
        zip_bytes = download(url)
        content_hash = sha256_bytes(zip_bytes)
        print(f"  {len(zip_bytes) / 1e6:.2f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/sfos_afr.zip"
        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=zip_bytes,
                mime_type="application/zip",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/zip",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                f"AFR sheet; Account='{TOPLINE_ACCOUNT}' (TOTAL GENERAL "
                f"FUND EXPENDITURES); group by AgencyID; "
                f"AgencyID.replace('-','') == state_leaid suffix"
            ),
            notes=(
                f"FY{fiscal_year} NE SFOS AFR. Fetched via curl-cffi "
                f"chrome120 (no WAF actually present, but used "
                f"consistently with other extractors). General Fund "
                f"operating only; narrower than F-33 frame."
            ),
        )

        crosswalk = build_ne_crosswalk(client)
        print(f"  NE crosswalk: {len(crosswalk):,} state→NCES mappings")

        records = parse_ne_afr(zip_bytes)
        print(f"  AFR records (General Fund expenditures): {len(records):,}")

        no_match: list[str] = []
        for d in records:
            # AgencyID in file = '55-0001-000'; state_leaid suffix = '550001000'
            key = d["agency_id"].replace("-", "")
            district = crosswalk.get(key)
            if district is None:
                no_match.append(d["agency_id"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            _, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched AgencyIDs (ESUs / coops / non-master): {len(no_match)}"
        )
        if no_match[:5]:
            print(f"  sample unmatched: {no_match[:8]}")

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2025)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
