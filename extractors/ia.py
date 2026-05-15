"""Iowa extractor — Iowa DE Certified Annual Report (CAR) workbook.

Source: https://educate.iowa.gov/pk-12/operation-support/business-finance/accounting-reporting/certified-annual-report
File: media/{nnnnn}/download — '{YYYY-YYYY} CAR data' XLSX. Iowa DE
      Bureau of Finance compiles district-submitted CARs into a
      single multi-sheet workbook (one sheet per fund × {presentation,
      data}). Data sheets are named *ExpData1, *ExpData2, etc.

What this gives us:
  - Per-district expenditure detail by Function × Object across all
    Iowa school district funds (General, Activity, Management,
    Nutrition, Library, plus Capital/Debt/Trust which we exclude).
    ~336 rows in each *ExpData sheet (325 districts + ~11 AEAs).

Topline definition:
  Sum of all numeric expenditure cells (cols 3..N per row) across
  the four core operating-fund data sheets:
    GenExpData1 (General Fund), ActExpData1 (Activity),
    MgmntExpData1 (Management), NutritionExpData1 (Nutrition).
  Excludes CapProjExpData1 (Capital), DebtExpData1, SAVEExpData1
  and PPELExpData1 (sales-tax + physical-plant capital), permanent
  trust, internal services, and AEA-only sheets. Aligned with F-33
  'current expenditures' frame.

Status: `actual` — post-CAR audited.

Crosswalk:
  Master state_leaid format: 'IA-{6-digit} 000' (e.g. 'IA-420009 000'
                              AGWSR Comm School District)
  CAR district column:        integer (e.g. 9)
  → state_leaid suffix last 4 chars (drop trailing ' 000') with
    leading zeros stripped == CAR district.
"""

from __future__ import annotations

import argparse
import io
import sys
import urllib.error
import urllib.request

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "ia"
STATE = "IA"
BUCKET = "ia"
SOURCE_PORTAL_URL = "https://educate.iowa.gov/pk-12/operation-support/business-finance/accounting-reporting/certified-annual-report"
PUBLISHER = "Iowa Department of Education (Bureau of Finance, Facilities, Operations & Transportation)"
DOCUMENT_TYPE = "iowa_de_car_xlsx"
TOPLINE_DEFINITION = (
    "Iowa DE Certified Annual Report — sum across the four core "
    "operating-fund data sheets {GenExpData1, ActExpData1, "
    "MgmntExpData1, NutritionExpData1} of all expenditure cells "
    "(cols 3..N per row), grouped by district number. Excludes "
    "Capital Projects, Debt, SAVE/PPEL (sales-tax + physical-plant "
    "capital), Permanent Trust, Internal Services, and AEA sheets. "
    "Aligned with F-33 'current expenditures' frame."
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

OPERATING_SHEETS = (
    "GenExpData1",
    "ActExpData1",
    "MgmntExpData1",
    "NutritionExpData1",
)

# Phase 7.5 — IA per-sheet → canonical category mapping for the
# universal-floor component pass. NutritionExpData1 is also part of
# topline (operating) but emits a food_service component as well.
_IA_NONOPERATING_TO_CATEGORY: dict[str, str] = {
    "CapProjExpData1": "capital_outlay",
    "SAVEExpData1": "capital_outlay",
    "PPELExpData1": "capital_outlay",
    "DebtExpData1": "debt_service",
}

KNOWN_FILE_URLS: dict[int, str] = {
    # FY24 = SY 2023-24 CAR data; published 2025.
    2024: "https://educate.iowa.gov/media/9108/download?inline",
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def _ia_iter_sheet(ws, header_label="district"):
    """Yield (district_int, header_row_tuple, data_row_tuple) for each
    data row in a CAR ExpData sheet. header_row_tuple is column headers
    aligned to data row indices (so r[i] and header[i] correspond)."""
    header_row_idx = None
    for i, r in enumerate(ws.iter_rows(values_only=True, max_row=8)):
        if r and len(r) > 1 and r[1] == header_label:
            header_row_idx = i + 1
            header = r
            break
    if header_row_idx is None:
        return
    for r in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not r or r[1] is None:
            continue
        try:
            d = int(r[1])
        except (TypeError, ValueError):
            continue
        yield d, header, r


def parse_ia(xlsx_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    totals: dict[int, float] = {}
    components: dict[int, dict[str, float]] = {}
    # Operating-sheet pass — topline + per-sheet canonical breakdown.
    for sheet_name in OPERATING_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  WARN: sheet {sheet_name!r} not found; skipping")
            continue
        ws = wb[sheet_name]
        for d, header, r in _ia_iter_sheet(ws):
            row_sum = 0.0
            inst = 0.0
            benefits = 0.0
            for idx in range(3, len(r)):
                v = r[idx]
                if v is None:
                    continue
                try:
                    f = float(v)
                except (TypeError, ValueError):
                    continue
                row_sum += f
                col_label = header[idx] if idx < len(header) else None
                if not col_label:
                    continue
                cl = str(col_label)
                # InstSal, InstBen, InstPurchServ, ... → instruction
                if cl.startswith("Inst"):
                    inst += f
                # *Ben columns (object 200 Employee Benefits)
                if cl.endswith("Ben"):
                    benefits += f
            if row_sum > 0:
                totals[d] = totals.get(d, 0.0) + row_sum
                if sheet_name == "NutritionExpData1":
                    components.setdefault(d, {}).setdefault("food_service", 0.0)
                    components[d]["food_service"] += row_sum
            if inst > 0:
                components.setdefault(d, {}).setdefault("instruction", 0.0)
                components[d]["instruction"] += inst
            if benefits > 0:
                components.setdefault(d, {}).setdefault("employee_benefits", 0.0)
                components[d]["employee_benefits"] += benefits

    # Non-operating capital + debt sheets — emit per-sheet sum as the
    # appropriate canonical category.
    for sheet_name, category in _IA_NONOPERATING_TO_CATEGORY.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for d, header, r in _ia_iter_sheet(ws):
            row_sum = 0.0
            for v in r[3:]:
                if v is None:
                    continue
                try:
                    row_sum += float(v)
                except (TypeError, ValueError):
                    continue
            if row_sum > 0:
                components.setdefault(d, {}).setdefault(category, 0.0)
                components[d][category] += row_sum

    return [
        {
            "code": str(d),
            "total_op_exp": v,
            "components": components.get(d, {}),
        }
        for d, v in totals.items()
        if v > 0
    ]


def build_ia_crosswalk(client: Client) -> dict[str, dict]:
    """Master state_leaid 'IA-{6-digit} 000' → district number = last 4
    digits of the 6-digit prefix, leading zeros stripped."""
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if not sl.startswith("IA-"):
            continue
        # 'IA-420009 000' -> '420009' -> last 4 -> '0009' -> '9'
        suffix = sl.removeprefix("IA-").split(" ")[0]
        if len(suffix) >= 4:
            code = suffix[-4:].lstrip("0") or "0"
            out[code] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"IA extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No Iowa CAR URL for fiscal_year={fiscal_year}; "
                "add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/', 1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.1f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/iowa_car.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheets {GenExpData1, ActExpData1, MgmntExpData1, "
                "NutritionExpData1}; sum cols 3..N per row; group by "
                "district col; match district == lstrip('0', state_leaid "
                "suffix last 4 digits)"
            ),
            notes=f"FY{fiscal_year} Iowa DE Certified Annual Report (CAR) workbook",
        )

        crosswalk = build_ia_crosswalk(client)
        print(f"  IA crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_data = parse_ia(xlsx_bytes)
        print(f"  CAR districts with operating expenditures: {len(district_data):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for d in district_data:
            district = crosswalk.get(d["code"])
            if district is None:
                no_match.append(d["code"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            components: list[ComponentInput] = []
            for category, amount in d.get("components", {}).items():
                if amount <= 0:
                    continue
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=(
                            f"Iowa CAR workbook: '{category}' = sum of "
                            f"per-sheet/per-column-prefix amounts. "
                            f"capital_outlay = CapProj+SAVE+PPEL sheets; "
                            f"debt_service = Debt sheet; food_service = "
                            f"Nutrition sheet; instruction = cols starting "
                            f"with 'Inst' in operating sheets; "
                            f"employee_benefits = cols ending in 'Ben'."
                        ),
                        line_or_cell_reference=(
                            f"district col 1={d['code']}; per "
                            f"_IA_NONOPERATING_TO_CATEGORY + Inst*/+*Ben prefix"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched CAR districts (AEAs/specialty): {len(no_match)}"
        )
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
