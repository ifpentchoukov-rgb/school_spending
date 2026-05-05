"""Washington extractor — OSPI F-196 10-Year Historical Data Detail.

Source: https://ospi.k12.wa.us/policy-funding/school-apportionment/
        school-publications/financial-reporting-summary
File: 10_year_f-196_data_2024-25.xlsx (latest as of 2026-05-05; OSPI
publishes a new 10-year file annually each fall after F-196 reconciliation).

What this gives us:
  - Per-district General Fund expenditures + revenues by source, program,
    activity, and object across 11 fiscal years (FY15-FY25 in the latest
    file). 327 LEAs covered.
  - WA fiscal year is Sept 1 – Aug 31 (NOT July-June like most states),
    so FY25 = SY 2024-25 ending Aug 31, 2025. F-196 published Dec 2025.

Topline definition:
  `EXP by District` sheet, last year column ('24-25') — General Fund
  total expenditures per district. This is WA's standard all-funds
  operating spend, aligned with F-33 'current expenditures' frame.

Status: `actual` — F-196 is the audited annual financial report.

Crosswalk:
  Master state_leaid format: 'WA-{5-digit-CCDDD}' (e.g. 'WA-17001' Seattle)
  F-196 column B:            5-digit zero-padded CCDDD
  → strip 'WA-'.
"""

from __future__ import annotations

import argparse
import sys
import urllib.error
import urllib.request
from io import BytesIO

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "wa"
STATE = "WA"
BUCKET = "wa"
SOURCE_PORTAL_URL = (
    "https://ospi.k12.wa.us/policy-funding/school-apportionment/"
    "school-publications/financial-reporting-summary"
)
PUBLISHER = "Washington Office of Superintendent of Public Instruction"
DOCUMENT_TYPE = "ospi_f196_10yr_xlsx"
TOPLINE_DEFINITION = (
    "OSPI F-196 10-year historical data, 'EXP by District' sheet, last "
    "year column — General Fund total expenditures per district"
)
USER_AGENT = "school-budget-tracker/0.1 (https://github.com/ifpentchoukov-rgb/school_spending)"

# OSPI publishes one 10-year file per FY release. The path includes the
# year's posting subdirectory (e.g. /2025-12/...). Hard-code the latest
# released file; add a new entry per year.
KNOWN_FILE_URLS: dict[int, str] = {
    2025: "https://ospi.k12.wa.us/sites/default/files/2025-12/10_year_f-196_data_2024-25.xlsx",
}

# In the EXP by District sheet, the last (most recent) year column is the
# 13th column (0-indexed 12 — covering 14-15 through 24-25 = 11 years).
LAST_YEAR_COL_IDX = 12
CCDDD_COL_IDX = 1
NAME_COL_IDX = 2


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def parse_exp_by_district(xlsx_bytes: bytes) -> list[dict]:
    """Return [{ccddd, name, total_exp}, ...] from the EXP by District sheet."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "EXP by District" not in wb.sheetnames:
        raise RuntimeError(
            f"'EXP by District' sheet missing; sheets={wb.sheetnames}"
        )
    ws = wb["EXP by District"]
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    for r in rows:
        if not r or not r[CCDDD_COL_IDX]:
            continue
        ccddd = str(r[CCDDD_COL_IDX]).strip()
        name = r[NAME_COL_IDX]
        # Skip header / total rows
        if not ccddd or ccddd in ("CCDDD", "") or not name or name in ("Grand Total", "Statewide", "District Name"):
            continue
        # Zero-pad to 5 digits (some codes might come through as int)
        ccddd = ccddd.zfill(5)
        try:
            total = float(r[LAST_YEAR_COL_IDX]) if r[LAST_YEAR_COL_IDX] is not None else None
        except (TypeError, ValueError):
            total = None
        if total is None or total <= 0:
            continue
        out.append({"ccddd": ccddd, "name": name, "total_exp": total})
    return out


def build_wa_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("WA-"):
            out[sl.removeprefix("WA-")] = r
    return out


def extract(*, fiscal_year: int = 2025, triggered_by: str = "manual") -> dict:
    print(f"WA extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No OSPI F-196 URL for fiscal_year={fiscal_year}; add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/',1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e3:.1f} KB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/f196_10yr.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'EXP by District'; match CCDDD column B == "
                "state_leaid suffix; topline = last year column "
                "(General Fund total expenditures)"
            ),
            notes=f"FY{fiscal_year} F-196 10-year historical data file",
        )

        crosswalk = build_wa_crosswalk(client)
        print(f"  WA crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_data = parse_exp_by_district(xlsx_bytes)
        print(f"  F-196 districts: {len(district_data):,}")

        no_match: list[str] = []
        for d in district_data:
            district = crosswalk.get(d["ccddd"])
            if district is None:
                no_match.append(d["ccddd"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            _, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched CCDDDs: {len(no_match)}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2025)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
