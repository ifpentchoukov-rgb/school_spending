"""Colorado actuals extractor — CDE Financial Transparency district data.

Per HB14-1292, every CO school district publishes its Annual Financial
Report by Jan 31 following the close of the FY. CDE consolidates these
into the 'Financial Transparency Disclosure' workbook each year.

Source URL:
  https://www.cde.state.co.us/cdefinance/ft_fy{YYYY}_distdatafile
  (returns the .xlsx; redirects via Content-Disposition header).

Network note:
  www.cde.state.co.us aggressively rate-limits non-browser HTTP clients
  ('rate-limited our IP'). We use `curl_cffi` with `impersonate='chrome120'`
  to mimic a real Chrome TLS handshake — that bypasses the WAF cleanly,
  but CDE will start refusing connections after ~10 requests in quick
  succession (per-IP throttle). The extractor retries with exponential
  backoff. `verify=False` because curl-impersonate doesn't pick up
  macOS's cert bundle automatically.

Topline definition:
  Sum of AMOUNT in sheet 'Org_Spending_Funding' where
  SPENDING_FUNDING='Spending' and ORG_ROLLUP in
  ('Learning Environment', 'Operations'). This excludes
  'Construction, Debt, Refinancing & Other' (capital + debt service)
  by ORG_ROLLUP filter. F-33 'current expenditures' frame.

Status: `actual` — Annual Financial Report (audited).

Crosswalk:
  Master state_leaid format: 'CO-{4-digit ORG_CODE}' (e.g. 'CO-0880'
                              School District No. 1 in Denver County)
  XLSX column ORG_CODE:       4-digit zero-padded
  → state_leaid suffix == ORG_CODE.

Charter coverage:
  Charters' financials roll up into their authorizing district's row;
  CDE does not break them out separately in this file.
"""

from __future__ import annotations

import argparse
import io
import sys
import time

import openpyxl
from curl_cffi import requests as curl_req
from curl_cffi.requests.exceptions import ConnectionError as CurlConnError
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    ComponentInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_components,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "co"
STATE = "CO"
BUCKET = "co"
SOURCE_PORTAL_URL = (
    "https://www.cde.state.co.us/schoolview/financialtransparency/homepage"
)
PUBLISHER = "Colorado Department of Education (Financial Transparency)"
DOCUMENT_TYPE = "cde_ft_distdatafile_xlsx"
TOPLINE_DEFINITION = (
    "CDE Financial Transparency Disclosure XLSX, sheet "
    "'Org_Spending_Funding'; sum AMOUNT where "
    "SPENDING_FUNDING='Spending' and ORG_ROLLUP in "
    "('Learning Environment', 'Operations'). Excludes "
    "'Construction, Debt, Refinancing & Other' by ORG_ROLLUP filter "
    "(= capital + debt service). F-33 'current expenditures' frame. "
    "Charters folded into their authorizing district's row."
)


# Annual CDE publication URLs — pinned per FY. CDE uses 'fy{YYYY}' where
# YYYY = the year FY ended (e.g. fy2024 = SY 2023-24 = our fiscal_year=2024).
KNOWN_FILE_URLS: dict[int, str] = {
    # FY24 (SY 2023-24) — published winter 2025
    2024: "https://www.cde.state.co.us/cdefinance/ft_fy2024_distdatafile",
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str, *, max_attempts: int = 5) -> bytes:
    """Fetch CDE file with exponential backoff. CDE's WAF can refuse
    connections under rate-limit even from a chrome120 impersonator;
    we wait progressively longer between retries."""
    last_err: Exception | None = None
    for attempt in range(max_attempts):
        try:
            r = curl_req.get(
                url,
                impersonate="chrome120",
                timeout=120,
                verify=False,
            )
            r.raise_for_status()
            return r.content
        except (CurlConnError, Exception) as e:
            last_err = e
            if attempt + 1 < max_attempts:
                wait = 30 * (2 ** attempt)  # 30, 60, 120, 240, 480 s
                print(
                    f"  ⏳ CDE fetch attempt {attempt + 1}/{max_attempts} "
                    f"failed ({type(e).__name__}); waiting {wait}s before retry..."
                )
                time.sleep(wait)
    raise RuntimeError(
        f"CDE fetch failed after {max_attempts} attempts. "
        f"Last error: {last_err!r}"
    )


# Phase 7.5 — CO Org_Explore_More (CATEGORY, ROLLUP) → canonical category.
# Each district's spending lands at SUB_ROLLUP='None' for each Rollup.
_CO_ROLLUP_TO_CATEGORY: dict[tuple[str, str], str] = {
    ("Learning Environment", "Instructional"): "instruction",
    ("Learning Environment", "District Administration"): "administration",
    ("Learning Environment", "School Administration"): "administration",
    ("Learning Environment", "Staff Support"): "support_services_instruction",
    ("Learning Environment", "Student Support"): "support_services_student",
    ("Operations", "Food Services"): "food_service",
    ("Operations", "Operations Maintenance"): "operations_maintenance",
    ("Operations", "Transportation"): "transportation",
    ("Operations", "Other Support"): "other",
    ("Construction, Debt, Refinancing & Other",
     "Construction Facilities Acquisitions And Construction Services"):
        "capital_outlay",
    ("Construction, Debt, Refinancing & Other", "Debt Services & Other Uses"):
        "debt_service",
}


def parse_co_ft(xlsx_bytes: bytes) -> list[dict]:
    """Return [{org_code, total_op_exp, components}] from the Financial
    Transparency XLSX. Topline = Org_Spending_Funding sum of Learning
    Environment + Operations. Components from Org_Explore_More keyed by
    (CATEGORY, ROLLUP) with SUB_ROLLUP='None'."""
    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), data_only=True, read_only=True
    )
    if "Org_Spending_Funding" not in wb.sheetnames:
        raise RuntimeError(
            f"Expected sheet 'Org_Spending_Funding' not found. "
            f"Sheets: {wb.sheetnames}"
        )

    # --- Topline pass (Org_Spending_Funding) ---
    ws = wb["Org_Spending_Funding"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    try:
        i_org = header.index("ORG_CODE")
        i_sf = header.index("SPENDING_FUNDING")
        i_roll = header.index("ORG_ROLLUP")
        i_amt = header.index("AMOUNT")
    except ValueError as e:
        raise RuntimeError(
            f"Missing expected column in Org_Spending_Funding: {e}"
        ) from e

    operating_rollups = {"Learning Environment", "Operations"}
    totals: dict[str, float] = {}
    for row in rows:
        if not row or row[i_org] is None:
            continue
        if row[i_sf] != "Spending":
            continue
        if row[i_roll] not in operating_rollups:
            continue
        amt = row[i_amt]
        if amt is None:
            continue
        try:
            amt_f = float(amt)
        except (TypeError, ValueError):
            continue
        org = str(row[i_org]).strip().zfill(4)
        if not org:
            continue
        totals[org] = totals.get(org, 0.0) + amt_f

    # --- Component pass (Org_Explore_More) ---
    components: dict[str, dict[str, float]] = {}
    if "Org_Explore_More" in wb.sheetnames:
        ws2 = wb["Org_Explore_More"]
        rows2 = ws2.iter_rows(values_only=True)
        header2 = next(rows2)
        try:
            i2_org = header2.index("ORG_CODE")
            i2_sf = header2.index("SPENDING_FUNDING")
            i2_cat = header2.index("CATEGORY")
            i2_roll = header2.index("ROLLUP")
            i2_sub = header2.index("SUB_ROLLUP")
            i2_amt = header2.index("AMOUNT")
        except ValueError:
            i2_org = None
        if i2_org is not None:
            for row in rows2:
                if not row or row[i2_org] is None:
                    continue
                if row[i2_sf] != "Spending":
                    continue
                cat = str(row[i2_cat]) if row[i2_cat] else ""
                roll = str(row[i2_roll]) if row[i2_roll] else ""
                sub = row[i2_sub]
                amt = row[i2_amt]
                if amt is None:
                    continue
                try:
                    amt_f = float(amt)
                except (TypeError, ValueError):
                    continue
                if amt_f <= 0:
                    continue
                org = str(row[i2_org]).strip().zfill(4)
                # Map (CATEGORY, ROLLUP) → canonical, SUB_ROLLUP must be
                # None (the rollup total slice).
                category = _CO_ROLLUP_TO_CATEGORY.get((cat, roll))
                if category and sub is None:
                    components.setdefault(org, {}).setdefault(category, 0.0)
                    components[org][category] += amt_f
                # employee_benefits = sum of ('Spending Overview', 'Benefits', X)
                # across all personnel-type sub_rollups (no None aggregate).
                if cat == "Spending Overview" and roll == "Benefits" and sub:
                    components.setdefault(org, {}).setdefault(
                        "employee_benefits", 0.0
                    )
                    components[org]["employee_benefits"] += amt_f

    return [
        {
            "org_code": k,
            "total_op_exp": v,
            "components": components.get(k, {}),
        }
        for k, v in totals.items()
        if v > 0
    ]


def build_co_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("CO-"):
            out[sl.removeprefix("CO-").strip().zfill(4)] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual",
            xlsx_path: str | None = None) -> dict:
    """If xlsx_path is given, parse that file directly (skip CDE download).
    Useful while CDE rate-limits our IP."""
    print(f"CO actuals extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No CO Financial Transparency URL for fiscal_year="
                f"{fiscal_year}; add to KNOWN_FILE_URLS."
            )

        if xlsx_path:
            print(f"  reading from {xlsx_path} (CDE rate-limit fallback)...")
            with open(xlsx_path, "rb") as f:
                xlsx_bytes = f.read()
        else:
            print(f"  downloading {url} (curl-cffi chrome120, with retries)...")
            xlsx_bytes = download(url)

        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.2f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/ft_distdatafile.xlsx"
        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'Org_Spending_Funding'; sum AMOUNT where "
                "SPENDING_FUNDING='Spending' and ORG_ROLLUP in "
                "('Learning Environment', 'Operations'); group by "
                "ORG_CODE; ORG_CODE == state_leaid suffix"
            ),
            notes=(
                f"FY{fiscal_year} CO CDE Financial Transparency district "
                f"data. Fetched via curl-cffi chrome120 (CDE WAF + IP "
                f"rate-limit). F-33 frame: Learning Environment + "
                f"Operations only; Construction/Debt excluded."
            ),
        )

        crosswalk = build_co_crosswalk(client)
        print(f"  CO crosswalk: {len(crosswalk):,} state→NCES mappings")

        records = parse_co_ft(xlsx_bytes)
        print(f"  CDE records (Spending operating): {len(records):,}")

        no_match: list[str] = []
        n_components_inserted = 0
        n_components_updated = 0
        n_components_unchanged = 0
        for d in records:
            district = crosswalk.get(d["org_code"])
            if district is None:
                no_match.append(d["org_code"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            event_id, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

            components: list[ComponentInput] = []
            for category, amount in d.get("components", {}).items():
                if amount <= 0:
                    continue
                components.append(
                    ComponentInput(
                        category=category,
                        amount=float(amount),
                        definition=(
                            f"CDE Financial Transparency 'Org_Explore_More' "
                            f"sheet: (CATEGORY, ROLLUP) mapping to '{category}' "
                            f"with SUB_ROLLUP=None (rollup total slice). For "
                            f"employee_benefits: sum of ('Spending Overview', "
                            f"'Benefits', *) across personnel-type sub-rollups."
                        ),
                        line_or_cell_reference=(
                            f"Sheet 'Org_Explore_More'; ORG_CODE={d['org_code']}; "
                            f"per _CO_ROLLUP_TO_CATEGORY"
                        ),
                    )
                )
            if components:
                ins, upd, unch = upsert_components(
                    client=client,
                    budget_event_id=event_id,
                    components=components,
                )
                n_components_inserted += ins
                n_components_updated += upd
                n_components_unchanged += unch

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched ORG_CODEs (BOCES / state aggregators / non-master): "
            f"{len(no_match)}"
        )
        if no_match[:5]:
            print(f"  sample unmatched: {no_match[:8]}")
        print(
            f"  components: inserted={n_components_inserted} "
            f"updated={n_components_updated} unchanged={n_components_unchanged}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    p.add_argument(
        "--xlsx-path",
        default=None,
        help="Local XLSX path (skip CDE download; used when CDE rate-limits)",
    )
    args = p.parse_args()
    extract(
        fiscal_year=args.fiscal_year,
        triggered_by=args.triggered_by,
        xlsx_path=args.xlsx_path,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
