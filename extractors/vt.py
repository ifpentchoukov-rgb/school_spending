"""Vermont extractor — VT AOE Cohort Spending by School Type XLSX.

Source: https://education.vermont.gov/data-and-reporting/financial-reports/pupil-spending
File: edu-FY{YY}-Cohort-Spending-by-School-Type-v{NN}-web.xlsx
      Vermont Agency of Education publishes annually with one row per
      school district plus rankings.

What this gives us:
  - Per-district FY{YY} Equalized Pupils, Budgets per Equalized Pupil,
    Education Spending per Equalized Pupil, plus rankings.

Topline definition:
  Sheet 'SpendData FY{YY}rpt': Equalized Pupils (col 5) × Education
  Spending per Equalized Pupil (col 11). Vermont's 'Education Spending'
  is the F-33-aligned operating expenditure measure (excludes capital
  and debt service).

Status: `actual` — post-AFR audited.

Crosswalk:
  Master state_leaid format: 'VT-{T###|U###}' (Town / Unified Union)
  XLSX LEA col:              'T101', 'U049', etc.
  → state_leaid suffix == LEA directly.
"""

from __future__ import annotations

import argparse
import io
import sys
import urllib.error
import urllib.request

import openpyxl
from supabase import Client

from extractors._base import (
    BudgetEventInput,
    Run,
    fetch_all,
    sha256_bytes,
    upload_source_document,
    upsert_budget_event_with_supersession,
    upsert_source_document_row,
)

EXTRACTOR_NAME = "vt"
STATE = "VT"
BUCKET = "vt"
SOURCE_PORTAL_URL = "https://education.vermont.gov/data-and-reporting/financial-reports/pupil-spending"
PUBLISHER = "Vermont Agency of Education"
DOCUMENT_TYPE = "vt_aoe_cohort_spending_xlsx"
TOPLINE_DEFINITION = (
    "VT AOE Cohort Spending by School Type XLSX, sheet 'SpendData FY{YY}"
    "rpt': Equalized Pupils × Education Spending per Equalized Pupil "
    "per district. Vermont 'Education Spending' is F-33-aligned current "
    "operating expenditure (excludes capital and debt service)."
)
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"

KNOWN_FILE_URLS: dict[int, str] = {
    2024: "https://education.vermont.gov/sites/aoe/files/documents/edu-FY24-Cohort-Spending-by-School-Type-v02-web.xlsx",
}


def file_url(fiscal_year: int) -> str | None:
    return KNOWN_FILE_URLS.get(fiscal_year)


def download(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def parse_vt(xlsx_bytes: bytes, fiscal_year: int) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    yy = fiscal_year - 2000
    sheet = f"SpendData FY{yy}rpt"
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"Expected sheet {sheet!r}; got {wb.sheetnames}")
    ws = wb[sheet]
    out: list[dict] = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r:
            continue
        lea = r[1]
        if not lea or not isinstance(lea, str):
            continue
        lea = lea.strip()
        if not (lea.startswith("T") or lea.startswith("U")):
            continue
        try:
            eq_pupils = float(r[5])
            spend_per_pupil = float(r[11])
        except (TypeError, ValueError):
            continue
        if eq_pupils <= 0 or spend_per_pupil <= 0:
            continue
        total = eq_pupils * spend_per_pupil
        out.append({"code": lea, "total_op_exp": total})
    return out


def build_vt_crosswalk(client: Client) -> dict[str, dict]:
    rows = fetch_all(
        client.table("districts")
        .select("leaid, lea_name, state_leaid")
        .eq("state_postal", STATE)
        .eq("is_operating_district", True)
    )
    out: dict[str, dict] = {}
    for r in rows:
        sl = r.get("state_leaid") or ""
        if sl.startswith("VT-"):
            out[sl.removeprefix("VT-")] = r
    return out


def extract(*, fiscal_year: int = 2024, triggered_by: str = "manual") -> dict:
    print(f"VT extract: fiscal_year={fiscal_year}")

    with Run(extractor_name=EXTRACTOR_NAME, triggered_by=triggered_by) as run:
        client = run.client

        url = file_url(fiscal_year)
        if not url:
            raise RuntimeError(
                f"No VT AOE URL for fiscal_year={fiscal_year}; add to KNOWN_FILE_URLS."
            )
        print(f"  downloading {url.rsplit('/', 1)[-1]}...")
        xlsx_bytes = download(url)
        content_hash = sha256_bytes(xlsx_bytes)
        print(f"  {len(xlsx_bytes) / 1e6:.2f} MB; sha256={content_hash[:12]}...")

        storage_relpath = f"fy{fiscal_year}/cohort_spending.xlsx"

        existing_src = (
            client.table("source_documents")
            .select("id")
            .eq("content_hash_sha256", content_hash)
            .execute()
        )
        if not existing_src.data:
            print(f"  uploading to {BUCKET}/{storage_relpath}...")
            upload_source_document(
                client=client,
                bucket=BUCKET,
                storage_path=storage_relpath,
                content=xlsx_bytes,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        src_id = upsert_source_document_row(
            client=client,
            content_hash=content_hash,
            source_url=url,
            storage_path=f"{BUCKET}/{storage_relpath}",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            publisher=PUBLISHER,
            document_type=DOCUMENT_TYPE,
            line_or_cell_reference=(
                "Sheet 'SpendData FY{YY}rpt'; col 5 (Eq Pupils) × col 11 "
                "(Ed Spending per Eq Pupil); match LEA col == state_leaid suffix"
            ),
            notes=f"FY{fiscal_year} VT AOE Cohort Spending by School Type",
        )

        crosswalk = build_vt_crosswalk(client)
        print(f"  VT crosswalk: {len(crosswalk):,} state→NCES mappings")

        district_data = parse_vt(xlsx_bytes, fiscal_year=fiscal_year)
        print(f"  VT AOE districts: {len(district_data):,}")

        no_match: list[str] = []
        for d in district_data:
            district = crosswalk.get(d["code"])
            if district is None:
                no_match.append(d["code"])
                continue

            event = BudgetEventInput(
                leaid=district["leaid"],
                fiscal_year=fiscal_year,
                status="actual",
                topline_amount=d["total_op_exp"],
                topline_definition=TOPLINE_DEFINITION,
                source_document_id=src_id,
                extraction_run_id=run.run_id,
            )
            _, changed = upsert_budget_event_with_supersession(
                client=client, event=event
            )
            run.records_extracted += 1
            if changed:
                run.records_changed += 1

        print(
            f"  inserted/changed={run.records_changed}/{run.records_extracted}; "
            f"unmatched VT codes: {len(no_match)}"
        )

    return {
        "fiscal_year": fiscal_year,
        "records_extracted": run.records_extracted,
        "records_changed": run.records_changed,
        "no_match_count": len(no_match),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--fiscal-year", type=int, default=2024)
    p.add_argument("--triggered-by", default="manual",
                   choices=["cron", "manual", "backfill"])
    args = p.parse_args()
    extract(fiscal_year=args.fiscal_year, triggered_by=args.triggered_by)
    return 0


if __name__ == "__main__":
    sys.exit(main())
